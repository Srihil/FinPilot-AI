from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from app.db.base import get_db
from app.auth.dependencies import get_current_user, require_admin_or_accountant
from app.models.user import User
from app.models.upload import Upload, UploadType, UploadStatus
from app.models.audit_log import AuditAction
from app.models.tally_connector import TallyConnector, ConnectorStatus
from app.models.tally_job import TallyIntegrationJob, TallyJobOperation
from app.models.tally_masters import TallyLedger, TallyStockGroup, TallyUnit, TallyGodown
from app.services.audit_service import audit_service
from app.services import ingestion_service as ingest
from app.core.config import settings
import json
import pandas as pd
import io, os, uuid
from datetime import datetime, timezone

router = APIRouter(prefix="/uploads", tags=["uploads"])

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_SIZE_BYTES = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024

COLUMN_MAPS = {
    "customers":    ["name", "email", "phone", "address", "city", "state", "gst_number", "payment_terms_days"],
    "vendors":      ["name", "email", "phone", "address", "city", "state", "gst_number", "payment_terms_days"],
    "products":     ["sku", "name", "category", "purchase_price", "selling_price", "tax_rate", "stock_quantity", "reorder_threshold"],
    "expenses":     ["title", "category", "expense_date", "amount", "tax_amount", "description"],
    "ledgers":      ["name", "parent_group", "opening_balance"],
    "stock_groups": ["name", "parent"],
    "stock_items":  ["name", "stock_group", "stock_category", "unit", "rate", "opening_qty", "hsn_code", "gst_rate"],
    "units":        ["name", "symbol", "decimal_places"],
    "godowns":      ["name", "parent"],
}

VALID_UPLOAD_TYPES = list(COLUMN_MAPS.keys())


def _validate_file(file: UploadFile) -> None:
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")


def _read_df(content: bytes, filename: str) -> pd.DataFrame:
    ext = os.path.splitext(filename)[1].lower()
    return pd.read_csv(io.BytesIO(content)) if ext == ".csv" else pd.read_excel(io.BytesIO(content))


def _tally_key(company_id, name: str) -> str:
    return f"{company_id}::{str(name).strip().lower()}"


def _active_connector(db: Session, company_id):
    return db.query(TallyConnector).filter(
        TallyConnector.company_id == company_id,
        TallyConnector.status == ConnectorStatus.ACTIVE,
    ).first()


def _queue_job(db: Session, company_id, connector_id, operation: TallyJobOperation, payload: dict, ikey: str) -> bool:
    if db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
        return False
    db.add(TallyIntegrationJob(
        company_id=company_id,
        connector_id=connector_id,
        operation=operation,
        payload=payload,
        idempotency_key=ikey,
    ))
    return True


# ─── CSV Upload ───────────────────────────────────────────────────────────────

@router.post("/csv")
async def upload_csv(
    file: UploadFile = File(...),
    upload_type: str = Form(...),
    sync_to_tally: bool = Form(False),
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    if upload_type not in VALID_UPLOAD_TYPES:
        raise HTTPException(status_code=400, detail=f"Unknown upload type. Allowed: {', '.join(VALID_UPLOAD_TYPES)}")

    _validate_file(file)
    content = await file.read()
    if len(content) > MAX_SIZE_BYTES:
        raise HTTPException(status_code=400, detail=f"File too large. Max {settings.MAX_UPLOAD_SIZE_MB}MB")

    try:
        ut = UploadType(upload_type)
    except ValueError:
        ut = None

    record = Upload(
        company_id=current_user.company_id,
        uploaded_by=current_user.id,
        filename=f"{uuid.uuid4()}_{file.filename}",
        original_filename=file.filename,
        file_type=os.path.splitext(file.filename or "")[1].lower().strip("."),
        upload_type=ut,
        status=UploadStatus.PROCESSING,
    )
    db.add(record)
    db.commit()

    try:
        df = _read_df(content, file.filename)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        valid_rows, invalid_rows = [], []

        for idx, row in df.iterrows():
            errs = []
            name = str(row.get("name", "")).strip()
            title = str(row.get("title", "")).strip()

            if upload_type in ("customers", "vendors", "products", "ledgers", "stock_groups", "units", "godowns") and not name:
                errs.append("Name is required")
            if upload_type == "expenses" and not title:
                errs.append("Title is required")

            if errs:
                invalid_rows.append({"row": idx + 2, "field": "name", "message": "; ".join(errs)})
            else:
                valid_rows.append(row.to_dict())

        committed = 0
        tally_queued = 0
        now = datetime.now(timezone.utc)
        connector = _active_connector(db, current_user.company_id)

        # ── Customers ────────────────────────────────────────────────────────
        if upload_type == "customers":
            from app.models.customer import Customer
            for row in valid_rows:
                db.add(Customer(
                    company_id=current_user.company_id,
                    name=str(row.get("name", "")).strip(),
                    email=str(row.get("email", "")) or None,
                    phone=str(row.get("phone", "")) or None,
                    address=str(row.get("address", "")) or None,
                    city=str(row.get("city", "")) or None,
                    state=str(row.get("state", "")) or None,
                    gst_number=str(row.get("gst_number", "")) or None,
                    is_active=True,
                ))
                committed += 1

        # ── Vendors ──────────────────────────────────────────────────────────
        elif upload_type == "vendors":
            from app.models.vendor import Vendor
            for row in valid_rows:
                db.add(Vendor(
                    company_id=current_user.company_id,
                    name=str(row.get("name", "")).strip(),
                    email=str(row.get("email", "")) or None,
                    phone=str(row.get("phone", "")) or None,
                    gst_number=str(row.get("gst_number", "")) or None,
                    is_active=True,
                ))
                committed += 1

        # ── Products ─────────────────────────────────────────────────────────
        elif upload_type == "products":
            from app.models.product import Product
            for row in valid_rows:
                db.add(Product(
                    company_id=current_user.company_id,
                    name=str(row.get("name", "")).strip(),
                    sku=str(row.get("sku", "")) or None,
                    category=str(row.get("category", "")) or None,
                    purchase_price=float(row.get("purchase_price") or 0),
                    selling_price=float(row.get("selling_price") or 0),
                    tax_rate=float(row.get("tax_rate") or 18),
                    stock_quantity=float(row.get("stock_quantity") or 0),
                    reorder_threshold=float(row.get("reorder_threshold") or 10),
                ))
                committed += 1

        # ── Expenses ─────────────────────────────────────────────────────────
        elif upload_type == "expenses":
            from app.models.expense import Expense, ExpenseStatus
            for row in valid_rows:
                raw_date = row.get("expense_date")
                try:
                    exp_date = pd.to_datetime(raw_date).to_pydatetime().replace(tzinfo=timezone.utc) if raw_date else now
                except Exception:
                    exp_date = now
                db.add(Expense(
                    company_id=current_user.company_id,
                    title=str(row.get("title", "")).strip(),
                    category=str(row.get("category", "General")) or "General",
                    expense_date=exp_date,
                    amount=float(row.get("amount") or 0),
                    tax_amount=float(row.get("tax_amount") or 0),
                    total_amount=float(row.get("amount") or 0),
                    currency="INR",
                    status=ExpenseStatus.DRAFT,
                ))
                committed += 1

        # ── Ledgers ──────────────────────────────────────────────────────────
        elif upload_type == "ledgers":
            for row in valid_rows:
                name = str(row.get("name", "")).strip()
                key = _tally_key(current_user.company_id, name)
                exists = db.query(TallyLedger).filter(
                    TallyLedger.company_id == current_user.company_id,
                    TallyLedger.tally_key == key,
                ).first()
                if exists:
                    invalid_rows.append({"row": committed + 2, "field": "name", "message": f"Ledger '{name}' already exists"})
                    continue
                parent_group = str(row.get("parent_group", "Sundry Debtors")).strip() or "Sundry Debtors"
                opening_balance = float(row.get("opening_balance") or 0)
                ledger = TallyLedger(
                    company_id=current_user.company_id, name=name,
                    parent_group=parent_group, opening_balance=opening_balance,
                    tally_key=key, source="finpilot", tally_sync_status="pending", is_active=True,
                )
                db.add(ledger)
                db.flush()
                committed += 1
                if connector and sync_to_tally:
                    ikey = f"bulk_ledger::{key}"
                    if _queue_job(db, current_user.company_id, connector.id, TallyJobOperation.CREATE_LEDGER,
                                  {"name": name, "group": parent_group, "opening_balance": str(int(opening_balance))}, ikey):
                        tally_queued += 1

        # ── Stock Groups ─────────────────────────────────────────────────────
        elif upload_type == "stock_groups":
            for row in valid_rows:
                name = str(row.get("name", "")).strip()
                if name.lower() == "primary":
                    continue
                key = _tally_key(current_user.company_id, name)
                exists = db.query(TallyStockGroup).filter(
                    TallyStockGroup.company_id == current_user.company_id,
                    TallyStockGroup.tally_key == key,
                ).first()
                if exists:
                    invalid_rows.append({"row": committed + 2, "field": "name", "message": f"Stock group '{name}' already exists"})
                    continue
                parent = str(row.get("parent", "")).strip() or None
                if parent and parent.lower() == "primary":
                    parent = None
                sg = TallyStockGroup(
                    company_id=current_user.company_id, name=name, parent=parent,
                    tally_key=key, source="finpilot", tally_sync_status="pending", is_active=True,
                )
                db.add(sg)
                db.flush()
                committed += 1
                if connector and sync_to_tally:
                    payload = {"name": name}
                    if parent:
                        payload["parent"] = parent
                    ikey = f"bulk_sg::{key}"
                    if _queue_job(db, current_user.company_id, connector.id, TallyJobOperation.CREATE_STOCK_GROUP, payload, ikey):
                        tally_queued += 1

        # ── Units ────────────────────────────────────────────────────────────
        elif upload_type == "units":
            for row in valid_rows:
                name = str(row.get("name", "")).strip()
                key = _tally_key(current_user.company_id, name)
                exists = db.query(TallyUnit).filter(
                    TallyUnit.company_id == current_user.company_id,
                    TallyUnit.tally_key == key,
                ).first()
                if exists:
                    invalid_rows.append({"row": committed + 2, "field": "name", "message": f"Unit '{name}' already exists"})
                    continue
                symbol = str(row.get("symbol", name)).strip().replace(" ", "")[:8] or name[:8]
                decimals = int(float(row.get("decimal_places") or 0))
                unit = TallyUnit(
                    company_id=current_user.company_id, name=name, symbol=symbol,
                    decimal_places=decimals, unit_type="simple",
                    tally_key=key, source="finpilot", tally_sync_status="pending", is_active=True,
                )
                db.add(unit)
                db.flush()
                committed += 1
                if connector and sync_to_tally:
                    ikey = f"bulk_unit::{key}"
                    if _queue_job(db, current_user.company_id, connector.id, TallyJobOperation.CREATE_UNIT,
                                  {"name": name, "symbol": symbol, "decimal_places": str(decimals)}, ikey):
                        tally_queued += 1

        # ── Godowns ──────────────────────────────────────────────────────────
        elif upload_type == "godowns":
            for row in valid_rows:
                name = str(row.get("name", "")).strip()
                key = _tally_key(current_user.company_id, name)
                exists = db.query(TallyGodown).filter(
                    TallyGodown.company_id == current_user.company_id,
                    TallyGodown.tally_key == key,
                ).first()
                if exists:
                    invalid_rows.append({"row": committed + 2, "field": "name", "message": f"Godown '{name}' already exists"})
                    continue
                parent = str(row.get("parent", "")).strip() or None
                godown = TallyGodown(
                    company_id=current_user.company_id, name=name, parent=parent,
                    tally_key=key, source="finpilot", tally_sync_status="pending", is_active=True,
                )
                db.add(godown)
                db.flush()
                committed += 1
                if connector and sync_to_tally:
                    payload = {"name": name}
                    if parent:
                        payload["parent"] = parent
                    ikey = f"bulk_godown::{key}"
                    if _queue_job(db, current_user.company_id, connector.id, TallyJobOperation.CREATE_GODOWN, payload, ikey):
                        tally_queued += 1

        db.commit()

        record.status = UploadStatus.COMPLETED if not invalid_rows else UploadStatus.PARTIAL
        record.total_rows = len(df)
        record.valid_rows = len(valid_rows)
        record.invalid_rows = len(invalid_rows)
        record.imported_rows = committed
        record.error_summary = {"errors": invalid_rows[:50]} if invalid_rows else None
        record.completed_at = now
        db.commit()

        audit_service.log(db, current_user.company_id, current_user.id, AuditAction.UPLOAD,
                          entity_type="upload", entity_id=record.id,
                          description=f"Bulk upload {file.filename}: {committed} {upload_type} imported")

        return {
            "upload_id": str(record.id),
            "total_rows": len(df),
            "valid_rows": len(valid_rows),
            "invalid_rows": len(invalid_rows),
            "imported_rows": committed,
            "tally_queued": tally_queued,
            "status": record.status.value,
            "errors": [{"row": e.get("row"), "field": e.get("field"), "message": e.get("message")} for e in invalid_rows[:20]],
            "duplicate_rows": 0,
            "columns_detected": list(df.columns),
        }

    except HTTPException:
        raise
    except Exception as e:
        record.status = UploadStatus.FAILED
        db.commit()
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


# ─── PDF Invoice ──────────────────────────────────────────────────────────────

@router.post("/pdf-invoice")
async def upload_pdf_invoice(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are supported")

    content = await file.read()
    if len(content) > MAX_SIZE_BYTES:
        raise HTTPException(status_code=400, detail=f"File too large. Max {settings.MAX_UPLOAD_SIZE_MB}MB")

    from app.core.config import settings as cfg

    # Try AI extraction if provider configured
    extracted = {
        "vendor_name": "", "invoice_number": "", "invoice_date": "",
        "due_date": "", "subtotal": 0, "tax_amount": 0, "total_amount": 0,
        "line_items": [], "notes": "",
    }

    record = Upload(
        company_id=current_user.company_id,
        uploaded_by=current_user.id,
        filename=f"{uuid.uuid4()}_{file.filename}",
        original_filename=file.filename,
        file_type="pdf",
        upload_type=UploadType.INVOICE_PDF,
        status=UploadStatus.COMPLETED,
        extracted_data=extracted,
        completed_at=datetime.now(timezone.utc),
    )
    db.add(record)
    db.commit()

    audit_service.log(db, current_user.company_id, current_user.id, AuditAction.UPLOAD,
                      entity_type="upload", entity_id=record.id,
                      description=f"PDF invoice uploaded: {file.filename}")

    return {"upload_id": str(record.id), "extracted": extracted, "filename": file.filename}


# ─── Template download ────────────────────────────────────────────────────────

@router.get("/template/{upload_type}")
def download_template(upload_type: str):
    columns = COLUMN_MAPS.get(upload_type)
    if not columns:
        raise HTTPException(status_code=400, detail=f"Unknown upload type. Allowed: {', '.join(COLUMN_MAPS.keys())}")

    # Add a sample row so users know the format
    sample_data: dict[str, list] = {col: [] for col in columns}
    if upload_type == "customers":
        sample_data = {"name": ["Acme Corp"], "email": ["contact@acme.com"], "phone": ["9876543210"],
                       "address": ["123 MG Road"], "city": ["Mumbai"], "state": ["Maharashtra"],
                       "gst_number": ["27AAAAA0000A1Z5"], "payment_terms_days": [30]}
    elif upload_type == "vendors":
        sample_data = {"name": ["Tata Steel"], "email": ["accounts@tata.com"], "phone": ["9876543210"],
                       "address": ["Jamshedpur"], "city": ["Jamshedpur"], "state": ["Jharkhand"],
                       "gst_number": ["20AAAAA0000A1Z5"], "payment_terms_days": [45]}
    elif upload_type == "products":
        sample_data = {"sku": ["SKU-001"], "name": ["Widget A"], "category": ["Electronics"],
                       "purchase_price": [500], "selling_price": [750], "tax_rate": [18],
                       "stock_quantity": [100], "reorder_threshold": [10]}
    elif upload_type == "expenses":
        sample_data = {"title": ["Office Rent"], "category": ["Rent"], "expense_date": ["2026-08-01"],
                       "amount": [50000], "tax_amount": [0], "description": ["Monthly office rent"]}
    elif upload_type == "ledgers":
        sample_data = {"name": ["ABC Traders"], "parent_group": ["Sundry Debtors"], "opening_balance": [0]}
    elif upload_type == "stock_groups":
        sample_data = {"name": ["Electronics"], "parent": [""]}
    elif upload_type == "stock_items":
        sample_data = {
            "name":           ["Samsung Galaxy S24", "Dell Laptop 15in", "Office Chair Ergonomic"],
            "stock_group":    ["Electronics",        "Electronics",      "Furniture"],
            "stock_category": ["",                   "",                 ""],
            "unit":           ["Nos",                "Nos",              "Nos"],
            "rate":           [25000,                55000,              8500],
            "opening_qty":    [10,                   5,                  20],
            "hsn_code":       ["8517",               "8471",             "9401"],
            "gst_rate":       [18,                   18,                 18],
        }
    elif upload_type == "units":
        sample_data = {"name": ["Pieces"], "symbol": ["Pcs"], "decimal_places": [0]}
    elif upload_type == "godowns":
        sample_data = {"name": ["Main Warehouse"], "parent": [""]}

    df = pd.DataFrame(sample_data)
    buffer = io.BytesIO()
    df.to_csv(buffer, index=False)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={upload_type}_template.csv"}
    )


# ─── Smart Ingestion pipeline ─────────────────────────────────────────────────

INGEST_ALLOWED_EXT = {".csv", ".xlsx", ".xls", ".json"}
INGEST_MAX_BYTES = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024


@router.post("/ingest/parse")
async def ingest_parse(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    """
    Step 1 — Parse uploaded file and auto-detect entity type + column mapping.
    Returns metadata including detected type, confidence, column mapping suggestions,
    and the first 5 sample rows for the UI to display.
    """
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in INGEST_ALLOWED_EXT:
        raise HTTPException(400, f"Unsupported file. Allowed: {', '.join(INGEST_ALLOWED_EXT)}")

    content = await file.read()
    if len(content) > INGEST_MAX_BYTES:
        raise HTTPException(400, f"File too large. Max {settings.MAX_UPLOAD_SIZE_MB}MB")

    try:
        parsed = ingest.parse_file(content, file.filename)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    except Exception as exc:
        raise HTTPException(500, f"Failed to parse file: {exc}")

    try:
        fingerprint = ingest.header_fingerprint(parsed.columns)

        # Check cache first — may skip detection entirely
        cached = ingest._get_cached_mapping(current_user.company_id, fingerprint, db)
        if cached:
            detection = ingest.DetectionResult(
                entity_type=cached['entity_type'],
                subtype='',
                confidence='high',
                score=999,
                from_cache=True,
            )
            mapping = cached['column_mapping']
        else:
            detection = ingest.detect_entity_type(parsed.columns, parsed.rows[:3], current_user.company_id, db)
            mapping = ingest.suggest_column_mapping(parsed.columns, detection.entity_type)

        # Determine upload_type for the record
        ut_str = ingest.ENTITY_TO_UPLOAD_TYPE.get(detection.entity_type)
        try:
            ut = UploadType(ut_str) if ut_str else None
        except ValueError:
            ut = None

        # Create upload record and store pending rows
        record = Upload(
            company_id=current_user.company_id,
            uploaded_by=current_user.id,
            filename=f"{uuid.uuid4()}_{file.filename}",
            original_filename=file.filename,
            file_size=len(content),
            file_type=ext.lstrip('.'),
            upload_type=ut,
            status=UploadStatus.PENDING,
            detected_entity_type=detection.entity_type,
            entity_confidence=detection.confidence,
            entity_subtype=detection.subtype,
            column_mapping=mapping,
            header_pattern_key=fingerprint,
            total_rows=parsed.total,
            pending_rows=[
                {k: (v if not hasattr(v, 'isoformat') else v.isoformat()) for k, v in row.items()}
                for row in parsed.rows
            ],
        )
        db.add(record)
        db.commit()
        db.refresh(record)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(500, f"Failed to process file: {exc}")

    return {
        "upload_id": str(record.id),
        "detected_entity_type": detection.entity_type,
        "entity_subtype": detection.subtype,
        "entity_confidence": detection.confidence,
        "from_cache": detection.from_cache,
        "file_columns": parsed.columns,
        "mapping_suggestions": mapping,
        "schema_fields": list((ingest.SCHEMA_FIELDS.get(detection.entity_type) or {}).keys()),
        "sample_rows": parsed.rows[:5],
        "total_rows": parsed.total,
        "file_format": parsed.file_format,
    }


@router.post("/ingest/{upload_id}/validate")
def ingest_validate(
    upload_id: str,
    body: dict,
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    """
    Step 2 — Validate all parsed rows with user-confirmed entity type + column mapping.
    Returns per-row status (valid / warning / error) and a summary count.
    Stores row_report on the upload record for the commit step.
    """
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    entity_type: str = body.get("entity_type") or record.detected_entity_type or ""
    column_mapping: dict = body.get("column_mapping") or record.column_mapping or {}

    if not entity_type or entity_type not in ingest.SCHEMA_FIELDS:
        raise HTTPException(400, f"Unknown entity type '{entity_type}'")

    rows = record.pending_rows or []
    if not rows:
        raise HTTPException(400, "No parsed rows found — re-parse the file")

    row_results = ingest.validate_rows(rows, entity_type, column_mapping, current_user.company_id, db)

    summary = {"valid": 0, "warnings": 0, "errors": 0, "total": len(row_results)}

    def _safe_mapped(m: dict) -> dict:
        return {k: (v if v is None or isinstance(v, (int, float, bool, str)) else str(v))
                for k, v in m.items()}

    # Count all non-error rows (including new status variants) for summary
    for r in row_results:
        s = r.status
        if s == 'error':
            summary['errors'] += 1
        elif s in ('warning', 'duplicate_fuzzy', 'ref_similar'):
            summary['warnings'] += 1
        elif s in ('duplicate_exact',):
            summary['errors'] += 1  # blocked by default
        else:
            summary['valid'] += 1  # new, ref_missing (resolvable)

    # Persist confirmed mapping + full row report (includes new fields)
    record.detected_entity_type = entity_type
    record.column_mapping = column_mapping
    record.row_report = [
        {
            "row_id": r.row_id,
            "status": r.status,
            "errors": r.errors,
            "warnings": r.warnings,
            "mapped": _safe_mapped(r.mapped),
            "duplicate_info": r.duplicate_info,
            "ref_issues": r.ref_issues,
            "check_default": r.check_default,
            "force_import": r.force_import,
        }
        for r in row_results
    ]
    record.valid_rows = summary['valid'] + summary['warnings']
    record.invalid_rows = summary['errors']
    record.status = UploadStatus.PROCESSING
    db.commit()

    return {
        "upload_id": str(record.id),
        "summary": summary,
        "rows": [
            {
                "row_id": r.row_id,
                "status": r.status,
                "errors": r.errors,
                "warnings": r.warnings,
                "mapped": r.mapped,
                "duplicate_info": r.duplicate_info,
                "ref_issues": r.ref_issues,
                "check_default": r.check_default,
                "force_import": r.force_import,
            }
            for r in row_results
        ],
    }


@router.post("/ingest/{upload_id}/commit")
def ingest_commit(
    upload_id: str,
    body: dict,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    """
    Step 3 — Commit selected rows.
    Processes only the rows whose row_id appears in selected_row_ids and
    whose status is not 'error'. Runs as a FastAPI background task so the
    response is immediate; poll /ingest/{id}/status for progress.
    On completion, confirmed column mapping is saved to cache for future uploads.
    """
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    selected_ids: list[int] = body.get("selected_row_ids", [])
    sync_to_tally: bool = bool(body.get("sync_to_tally", False))

    if not selected_ids:
        raise HTTPException(400, "No rows selected for import")
    if not record.row_report:
        raise HTTPException(400, "Rows not validated — call /validate first")

    # Snapshot IDs for background task before resetting progress
    record.status = UploadStatus.PROCESSING
    record.imported_rows = 0
    db.commit()

    company_id = current_user.company_id
    user_id = current_user.id
    entity_type = record.detected_entity_type
    col_map = record.column_mapping or {}
    fingerprint = record.header_pattern_key

    def _run_commit():
        import logging
        from app.db.base import SessionLocal
        from app.models.upload_row import UploadRow, UploadRowStatus
        from app.services.ingestion_service import VOUCHER_SUBTYPE_OP
        from datetime import datetime, timezone as tz
        log = logging.getLogger(__name__)
        bg_db = SessionLocal()
        try:
            # ── Step 1: Commit entity rows to FinPilot DB ─────────────────────
            result = ingest.commit_rows(
                upload_record_id=upload_id,
                selected_row_ids=selected_ids,
                sync_to_tally=sync_to_tally,
                company_id=company_id,
                user_id=user_id,
                db=bg_db,
            )
            if result.imported > 0 and entity_type and fingerprint:
                ingest.save_mapping_cache(company_id, fingerprint, entity_type, col_map, bg_db)

            # Reload upload record (commit_rows did multiple commits)
            upload_rec = bg_db.query(Upload).filter(Upload.id == upload_id).first()
            committed_ids: set = set((upload_rec.commit_summary or {}).get("committed_row_ids", []))
            error_by_row: dict = {e["row_id"]: e["message"] for e in (upload_rec.commit_summary or {}).get("errors", [])}
            row_report: list = upload_rec.row_report or []
            now_dt = datetime.now(tz.utc)

            log.info(f"[import] entity={entity_type} imported={result.imported} committed_ids={len(committed_ids)} sync={sync_to_tally}")

            # ── Step 2: Queue Tally jobs (CRITICAL — own try/except) ──────────
            tally_queued = 0
            tally_job_ids: dict = {}  # rid → job UUID (for linking UploadRows later)

            if sync_to_tally and result.imported > 0:
                try:
                    connector = bg_db.query(TallyConnector).filter(
                        TallyConnector.company_id == company_id,
                        TallyConnector.status == ConnectorStatus.ACTIVE,
                    ).first()
                    log.info(f"[import] connector={'found' if connector else 'NONE'}")

                    if connector:
                        batch_id = uuid.uuid4()
                        for row in row_report:
                            rid = row.get("row_id")
                            if rid not in committed_ids:
                                continue
                            mapped = row.get("mapped", {})
                            name = str(mapped.get("name") or "").strip()
                            job_obj = None

                            if entity_type == "Ledger" and name:
                                group = str(mapped.get("parent_group") or "Sundry Debtors").strip()
                                ob = str(int(float(mapped.get("opening_balance") or 0)))
                                ikey = f"bulk_ledger::{upload_id}::{rid}"
                                if not bg_db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
                                    job_obj = TallyIntegrationJob(
                                        company_id=company_id, connector_id=connector.id,
                                        operation=TallyJobOperation.CREATE_LEDGER,
                                        payload={"name": name, "group": group, "opening_balance": ob},
                                        idempotency_key=ikey, batch_id=batch_id)

                            elif entity_type == "Stock Item" and name:
                                pl: dict = {"name": name}
                                if mapped.get("stock_group"): pl["stock_group"] = str(mapped["stock_group"])
                                if mapped.get("unit"):         pl["unit"]        = str(mapped["unit"])
                                if mapped.get("rate") is not None: pl["rate"]   = str(mapped["rate"])
                                ikey = f"bulk_si::{upload_id}::{rid}"
                                if not bg_db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
                                    job_obj = TallyIntegrationJob(
                                        company_id=company_id, connector_id=connector.id,
                                        operation=TallyJobOperation.CREATE_STOCK_ITEM,
                                        payload=pl, idempotency_key=ikey, batch_id=batch_id)

                            elif entity_type == "Stock Group" and name:
                                pl = {"name": name}
                                if mapped.get("parent"): pl["parent"] = str(mapped["parent"])
                                ikey = f"bulk_sg::{upload_id}::{rid}"
                                if not bg_db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
                                    job_obj = TallyIntegrationJob(
                                        company_id=company_id, connector_id=connector.id,
                                        operation=TallyJobOperation.CREATE_STOCK_GROUP,
                                        payload=pl, idempotency_key=ikey, batch_id=batch_id)

                            elif entity_type == "Voucher":
                                vtype_raw = str(mapped.get("voucher_type") or "").strip()
                                vtype = vtype_raw.title().removesuffix(" Voucher").strip()
                                op_v = VOUCHER_SUBTYPE_OP.get(vtype)
                                if op_v:
                                    pl = {k: str(v) for k, v in mapped.items() if v is not None}
                                    pl["voucher_type_name"] = vtype
                                    ikey = f"bulk_vch::{upload_id}::{rid}"
                                    if not bg_db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
                                        job_obj = TallyIntegrationJob(
                                            company_id=company_id, connector_id=connector.id,
                                            operation=op_v, payload=pl,
                                            idempotency_key=ikey, batch_id=batch_id)

                            if job_obj:
                                bg_db.add(job_obj)
                                bg_db.flush()
                                tally_queued += 1
                                tally_job_ids[rid] = job_obj.id

                        if tally_queued > 0 and upload_rec.commit_summary:
                            cs = dict(upload_rec.commit_summary)
                            cs["tally_queued"] = tally_queued
                            upload_rec.commit_summary = cs
                            flag_modified(upload_rec, "commit_summary")

                    bg_db.commit()  # commit Tally jobs
                    log.info(f"[import] tally_queued={tally_queued}")

                except Exception as tally_exc:
                    log.error(f"[import] Tally sync error: {tally_exc}", exc_info=True)
                    bg_db.rollback()
                    tally_queued = 0
                    tally_job_ids = {}

            # ── Step 3: Create UploadRows (non-critical — own try/except) ────
            try:
                for rid, msg in error_by_row.items():
                    row = next((r for r in row_report if r.get("row_id") == rid), {})
                    bg_db.add(UploadRow(
                        upload_id=upload_id, row_number=rid + 2,
                        entity_type=entity_type, status=UploadRowStatus.FAILED,
                        error_reason=msg, raw_data=row.get("mapped"),
                        created_at=now_dt, updated_at=now_dt,
                    ))
                for row in row_report:
                    rid = row.get("row_id")
                    if rid not in committed_ids:
                        continue
                    job_id = tally_job_ids.get(rid)
                    bg_db.add(UploadRow(
                        upload_id=upload_id, row_number=rid + 2,
                        entity_type=entity_type,
                        status=UploadRowStatus.PENDING if job_id else UploadRowStatus.IMPORTED,
                        tally_job_id=job_id,
                        raw_data=row.get("mapped"), created_at=now_dt, updated_at=now_dt,
                    ))
                bg_db.commit()
            except Exception as row_exc:
                log.warning(f"[import] UploadRow write failed (non-critical): {row_exc}")
                bg_db.rollback()

            # ── Step 4: Audit log ─────────────────────────────────────────────
            audit_service.log(
                bg_db, company_id, user_id, AuditAction.UPLOAD,
                entity_type="upload", entity_id=upload_id,
                description=f"Bulk import {entity_type}: {result.imported} rows, {tally_queued} Tally jobs",
            )
            bg_db.commit()

        except Exception as exc:
            log.error(f"[import] _run_commit fatal error: {exc}", exc_info=True)
        finally:
            bg_db.close()

    background_tasks.add_task(_run_commit)

    return {
        "upload_id": str(record.id),
        "started": True,
        "total_selected": len(selected_ids),
        "message": "Import started — poll /status for progress",
    }


@router.post("/ingest/{upload_id}/row-action")
def ingest_row_action(
    upload_id: str,
    body: dict,
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    """
    Apply a user-driven action to a single row in the row_report.

    Supported actions
    -----------------
    force_import   – mark a duplicate_exact row as force_import=True so it
                     gets committed even though an exact match exists in DB.
    use_existing   – overwrite a reference field with the user-chosen existing
                     value and re-resolve references for that row.
    dismiss_dup    – dismiss a duplicate_fuzzy warning without changing data
                     (row status becomes 'new' if no other issues remain).

    Returns the updated row dict.
    """
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    row_report: list[dict] = list(record.row_report or [])
    action: str = body.get("action", "")
    row_id: int = int(body.get("row_id", -1))
    field: str = body.get("field", "")
    resolved_value: str = body.get("resolved_value", "")

    # Find the target row
    row_idx = next((i for i, r in enumerate(row_report) if r["row_id"] == row_id), None)
    if row_idx is None:
        raise HTTPException(404, f"Row {row_id} not found in row_report")

    row = dict(row_report[row_idx])

    if action == "force_import":
        if row.get("status") != "duplicate_exact":
            raise HTTPException(400, "force_import is only valid for duplicate_exact rows")
        row["force_import"] = True
        row["check_default"] = True
        # Status becomes duplicate_exact_forced — UI can display differently
        row["status"] = "duplicate_exact_forced"

    elif action == "use_existing":
        if not field or not resolved_value:
            raise HTTPException(400, "field and resolved_value are required for use_existing")
        mapped = dict(row.get("mapped", {}))
        mapped[field] = resolved_value
        row["mapped"] = mapped
        # Re-resolve all references with the updated value
        entity_type = record.detected_entity_type or ""
        col_mapping = record.column_mapping or {}
        # Build updated ref_issues (quick re-check without full DB re-query)
        # The ref_issues for this specific field should now clear to 'ok'
        ref_issues = [
            ri for ri in row.get("ref_issues", [])
            if ri.get("field") != field
        ]
        row["ref_issues"] = ref_issues
        # Re-determine status
        from app.services.ingestion_service import _determine_final_status
        dup_info = row.get("duplicate_info")
        status, check_default = _determine_final_status(
            row.get("errors", []), row.get("warnings", []), dup_info, ref_issues
        )
        row["status"] = status
        row["check_default"] = check_default

    elif action == "dismiss_dup":
        row["duplicate_info"] = None
        from app.services.ingestion_service import _determine_final_status
        status, check_default = _determine_final_status(
            row.get("errors", []), row.get("warnings", []), None, row.get("ref_issues", [])
        )
        row["status"] = status
        row["check_default"] = check_default

    elif action == "update_mapped":
        # Bulk-update mapped field values and re-run full single-row validation
        updates: dict = body.get("updates", {})
        if not updates:
            raise HTTPException(400, "updates dict is required for update_mapped")
        mapped = dict(row.get("mapped", {}))
        mapped.update(updates)
        row["mapped"] = mapped

        # Full re-validation of this single row against DB
        entity_type = record.detected_entity_type or ""
        col_mapping = record.column_mapping or {}
        from app.services.ingestion_service import (
            validate_rows, _determine_final_status,
            _load_entity_records, _load_ref_records, ENTITY_REFERENCES, ENTITY_DB_TABLE,
            check_duplicate, resolve_entity_references, SCHEMA_FIELDS,
            _coerce_float, _coerce_date, _validate_ledger_fields,
            _validate_stock_item_fields, _validate_voucher, build_intra_file_index,
            TALLY_STANDARD_GROUPS,
        )

        # Re-validate this row's fields from scratch
        schema = SCHEMA_FIELDS.get(entity_type, {})
        errors: list = []
        warnings: list = []
        new_mapped: dict = {}

        for sf, meta in schema.items():
            val = mapped.get(sf)
            ft = meta["type"]
            if ft == "float":
                from app.services.ingestion_service import _coerce_float
                coerced = _coerce_float(val)
                if val is not None and val != "" and coerced is None:
                    errors.append({"field": sf, "message": f"'{val}' is not a valid number"})
                new_mapped[sf] = coerced
            elif ft == "date":
                from app.services.ingestion_service import _coerce_date
                coerced = _coerce_date(val)
                if val is not None and val != "" and coerced is None:
                    errors.append({"field": sf, "message": f"'{val}' is not a valid date"})
                new_mapped[sf] = coerced
            else:
                new_mapped[sf] = str(val) if val is not None else None
            if meta["required"] and (new_mapped.get(sf) is None or new_mapped.get(sf) == ""):
                errors.append({"field": sf, "message": f"'{sf}' is required"})

        if entity_type == "Ledger":
            company_groups: set = set()
            try:
                from sqlalchemy import text
                rows_g = db.execute(
                    text("SELECT name FROM tally_groups WHERE company_id = :cid AND is_active = true"),
                    {"cid": str(current_user.company_id)},
                ).fetchall()
                company_groups = {r[0].lower().strip() for r in rows_g} | TALLY_STANDARD_GROUPS
            except Exception:
                pass
            _validate_ledger_fields(new_mapped, errors, warnings, company_groups)
        elif entity_type == "Stock Item":
            _validate_stock_item_fields(new_mapped, warnings)
        elif entity_type == "Voucher":
            _validate_voucher(new_mapped, errors, warnings)

        # Duplicate check
        dup_info = None
        name_val = str(new_mapped.get("name", "")).lower().strip()
        if name_val and entity_type in ENTITY_DB_TABLE:
            entity_records = _load_entity_records(entity_type, current_user.company_id, db)
            dup_info = check_duplicate(name_val, entity_records)

        # Reference resolution
        ref_cache: dict = {}
        for ref_def in ENTITY_REFERENCES.get(entity_type, []):
            rt = ref_def["ref_type"]
            if rt not in ref_cache:
                ref_cache[rt] = _load_ref_records(rt, current_user.company_id, db)
        ref_issues = resolve_entity_references(entity_type, new_mapped, ref_cache, set())

        status, check_default = _determine_final_status(errors, warnings, dup_info, ref_issues)

        def _safe(v):
            return v if v is None or isinstance(v, (int, float, bool, str)) else str(v)

        row["mapped"] = {k: _safe(v) for k, v in new_mapped.items()}
        row["errors"] = errors
        row["warnings"] = warnings
        row["duplicate_info"] = dup_info
        row["ref_issues"] = ref_issues
        row["status"] = status
        row["check_default"] = check_default

    else:
        raise HTTPException(400, f"Unknown action '{action}'. Valid: force_import, use_existing, dismiss_dup, update_mapped")

    row_report[row_idx] = row
    record.row_report = row_report
    db.commit()

    return {"row": row}


@router.get("/sync-freshness")
def sync_freshness(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return last Tally sync timestamp for this company, and connector online status."""
    connector = db.query(TallyConnector).filter(
        TallyConnector.company_id == current_user.company_id,
        TallyConnector.status == ConnectorStatus.ACTIVE,
    ).first()

    return {
        "last_sync_at": connector.last_sync_at.isoformat() if (connector and connector.last_sync_at) else None,
        "connector_online": connector is not None,
        "connector_id": str(connector.id) if connector else None,
    }


# ─── Download remaining / full report ────────────────────────────────────────

def _build_file(rows: list[dict], columns: list[str], file_format: str,
                include_status: bool, buffer: io.BytesIO) -> str:
    """Generate CSV, XLSX, or JSON file; returns media_type string."""
    if file_format in ("xlsx", "xls"):
        import openpyxl
        from openpyxl.comments import Comment
        wb = openpyxl.Workbook()
        ws = wb.active
        # header
        hdr = columns + (["status", "error_reason"] if include_status else [])
        ws.append(hdr)
        for row in rows:
            vals = [row.get("data", {}).get(c) for c in columns]
            if include_status:
                vals += [row.get("status", ""), row.get("error_reason", "") or ""]
            ws.append(vals)
            # For remaining-rows XLSX, add comment to first cell of failed rows
            if not include_status and row.get("error_reason"):
                cell = ws.cell(row=ws.max_row, column=1)
                c = Comment(row["error_reason"], "FinPilot")
                c.width, c.height = 250, 60
                cell.comment = c
        wb.save(buffer)
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    elif file_format == "json":
        data = []
        for row in rows:
            entry = {k: v for k, v in row.get("data", {}).items()}
            if include_status:
                entry["_status"] = row.get("status", "")
                entry["_error_reason"] = row.get("error_reason", "") or ""
            data.append(entry)
        buffer.write(json.dumps(data, indent=2, default=str).encode())
        return "application/json"

    else:  # csv
        hdr = columns + (["status", "error_reason"] if include_status else [])
        import csv as csv_mod
        import io as io2
        text_buf = io2.StringIO()
        writer = csv_mod.DictWriter(text_buf, fieldnames=hdr, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            entry = {c: row.get("data", {}).get(c, "") for c in columns}
            if include_status:
                entry["status"] = row.get("status", "")
                entry["error_reason"] = row.get("error_reason", "") or ""
            writer.writerow(entry)
        buffer.write(text_buf.getvalue().encode("utf-8"))
        return "text/csv"


def _get_upload_rows_data(upload_id: str, company_id, db, statuses: list | None = None):
    """Return rows + column order + file format for an upload."""
    from app.models.upload_row import UploadRow
    q = db.query(UploadRow).filter(UploadRow.upload_id == upload_id)
    if statuses:
        q = q.filter(UploadRow.status.in_(statuses))
    ur_rows = q.order_by(UploadRow.row_number).all()

    record = db.query(Upload).filter(
        Upload.id == upload_id, Upload.company_id == company_id,
    ).first()

    # Derive column order from first pending_row
    pending = (record.pending_rows or []) if record else []
    columns = list(pending[0].keys()) if pending else []
    # Remap from pending_rows keys → column_mapping reverse lookup
    col_map = record.column_mapping or {} if record else {}
    reverse = {v: k for k, v in col_map.items()}  # schema_field → file_col
    # Build ordered column list matching original file
    ordered = []
    for file_col in (list(pending[0].keys()) if pending else []):
        if file_col not in ordered:
            ordered.append(file_col)
    if not ordered:
        # Fall back to schema fields
        ordered = list({sf for row in ur_rows for sf in (row.raw_data or {}).keys()})

    file_format = (record.file_type or "csv").lower() if record else "csv"

    rows_out = []
    for ur in ur_rows:
        raw = ur.raw_data or {}
        # Map schema_field values back to file column names
        file_row = {}
        for file_col in ordered:
            schema_field = col_map.get(file_col)
            file_row[file_col] = raw.get(schema_field) if schema_field else raw.get(file_col)
        rows_out.append({
            "data": file_row,
            "status": ur.status.value,
            "error_reason": ur.error_reason,
        })

    return rows_out, ordered, file_format


@router.get("/ingest/{upload_id}/remaining")
def download_remaining(
    upload_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download all non-imported rows (unselected + failed) in original file format, re-upload-ready."""
    from app.models.upload_row import UploadRow, UploadRowStatus
    record = db.query(Upload).filter(Upload.id == upload_id, Upload.company_id == current_user.company_id).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    committed_ids: set = set((record.commit_summary or {}).get("committed_row_ids", []))
    row_report: list = record.row_report or []

    rows_out: list = []
    columns: list = []

    if row_report:
        # Primary path: derive remaining rows from row_report.
        # "Remaining" = every row NOT in committed_ids (unselected + rows that errored
        # at DB level). This produces a clean re-upload-ready file.
        for row in row_report:
            if row.get("row_id") in committed_ids:
                continue
            mapped = row.get("mapped") or {}
            if not columns and mapped:
                columns = list(mapped.keys())
            errors = row.get("errors") or []
            rows_out.append({
                "data": mapped,
                "status": row.get("status", ""),
                "error_reason": "; ".join(e.get("message", "") for e in errors) if errors else None,
            })
    else:
        # Fallback for old imports that pre-date the row_report column:
        # only FAILED upload_rows remain to be re-imported (IMPORTED/PENDING are done).
        failed = db.query(UploadRow).filter(
            UploadRow.upload_id == upload_id,
            UploadRow.status == UploadRowStatus.FAILED,
        ).order_by(UploadRow.row_number).all()
        for ur in failed:
            raw = ur.raw_data or {}
            if not columns and raw:
                columns = list(raw.keys())
            rows_out.append({"data": raw, "status": "failed", "error_reason": ur.error_reason})

    if not rows_out:
        raise HTTPException(404, "No remaining rows — all rows were successfully imported")

    if not columns:
        columns = sorted({k for r in rows_out for k in r["data"].keys()})

    file_format = (record.file_type or "csv").lower()
    import datetime as _dt
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    orig = (record.original_filename or "upload").rsplit(".", 1)[0]
    filename = f"{orig}_remaining_{ts}.{file_format}"

    buf = io.BytesIO()
    media_type = _build_file(rows_out, columns, file_format, include_status=False, buffer=buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=media_type,
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/ingest/{upload_id}/report")
def download_report(
    upload_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download full audit report with status + error_reason columns."""
    record = db.query(Upload).filter(Upload.id == upload_id, Upload.company_id == current_user.company_id).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    rows, columns, file_format = _get_upload_rows_data(upload_id, current_user.company_id, db)

    if not rows:
        # Fallback for old imports without upload_rows records: build from row_report.
        committed_ids: set = set((record.commit_summary or {}).get("committed_row_ids", []))
        row_report: list = record.row_report or []
        columns = []
        rows = []
        for row in row_report:
            mapped = row.get("mapped") or {}
            if not columns and mapped:
                columns = list(mapped.keys())
            errors = row.get("errors") or []
            rid = row.get("row_id")
            status = "imported" if rid in committed_ids else row.get("status", "")
            rows.append({
                "data": mapped,
                "status": status,
                "error_reason": "; ".join(e.get("message", "") for e in errors) if errors else None,
            })
        if not columns:
            columns = sorted({k for r in rows for k in r["data"].keys()})
        file_format = (record.file_type or "csv").lower()

    if not rows:
        raise HTTPException(404, "No row data available")

    import datetime as _dt
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    orig = (record.original_filename or "upload").rsplit(".", 1)[0]
    ext = file_format if file_format != "xlsx" else "xlsx"
    filename = f"{orig}_report_{ts}.{ext}"

    buf = io.BytesIO()
    media_type = _build_file(rows, columns, file_format, include_status=True, buffer=buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=media_type,
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/ingest/{upload_id}/restore")
def ingest_restore(
    upload_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return enough data to rebuild the wizard UI after a page reload."""
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    from app.services.ingestion_service import SCHEMA_FIELDS

    # Derive which wizard step this upload is at
    if record.status == UploadStatus.PENDING:
        wizard_step = "mapping"
    elif record.status == UploadStatus.PROCESSING:
        if record.row_report and (record.imported_rows or 0) == 0 and not record.completed_at:
            wizard_step = "preview"
        else:
            wizard_step = "progress"
    elif record.status in (UploadStatus.COMPLETED, UploadStatus.PARTIAL, UploadStatus.FAILED):
        wizard_step = "done"
    else:
        wizard_step = "mapping"

    entity_type = record.detected_entity_type or ""
    result: dict = {
        "upload_id": str(record.id),
        "wizard_step": wizard_step,
        "entity_type": entity_type,
        "entity_confidence": record.entity_confidence or "low",
        "entity_subtype": record.entity_subtype or "",
        "column_mapping": record.column_mapping or {},
        "total_rows": record.total_rows or 0,
        "file_format": record.file_type or "csv",
    }

    # Reconstruct parse_result for the mapping step
    pending = record.pending_rows or []
    file_columns = list(pending[0].keys()) if pending else []
    result["parse_result"] = {
        "upload_id": str(record.id),
        "detected_entity_type": entity_type,
        "entity_subtype": record.entity_subtype or "",
        "entity_confidence": record.entity_confidence or "low",
        "from_cache": False,
        "file_columns": file_columns,
        "mapping_suggestions": record.column_mapping or {},
        "schema_fields": list((SCHEMA_FIELDS.get(entity_type) or {}).keys()),
        "sample_rows": pending[:5],
        "total_rows": record.total_rows or 0,
        "file_format": record.file_type or "csv",
    }

    # Reconstruct validate_result for the preview step
    if wizard_step == "preview" and record.row_report:
        rr = record.row_report
        summary = {"valid": 0, "warnings": 0, "errors": 0, "total": len(rr)}
        for r in rr:
            s = r.get("status", "")
            if s == "error":
                summary["errors"] += 1
            elif s in ("warning", "duplicate_fuzzy", "ref_similar"):
                summary["warnings"] += 1
            elif s == "duplicate_exact":
                summary["errors"] += 1
            else:
                summary["valid"] += 1
        result["validate_result"] = {
            "upload_id": str(record.id),
            "summary": summary,
            "rows": rr,
        }

    # Status data for progress/done step
    if wizard_step in ("progress", "done"):
        result["status_data"] = {
            "upload_id": str(record.id),
            "status": record.status.value,
            "entity_type": entity_type,
            "total_rows": record.total_rows or 0,
            "valid_rows": record.valid_rows or 0,
            "invalid_rows": record.invalid_rows or 0,
            "imported_rows": record.imported_rows or 0,
            "tally_queued": (record.commit_summary or {}).get("tally_queued", 0),
            "commit_summary": record.commit_summary,
            "completed_at": record.completed_at.isoformat() if record.completed_at else None,
        }

    return result


@router.get("/ingest/{upload_id}/status")
def ingest_status(
    upload_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Poll endpoint for commit progress. Returns current import counts and final summary."""
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")

    # Recover stale PROCESSING state (background task died > 10 min ago)
    if (record.status == UploadStatus.PROCESSING and record.created_at and
            (datetime.now(timezone.utc) - record.created_at).total_seconds() > 600 and
            not record.commit_summary):
        record.status = UploadStatus.FAILED
        db.commit()

    return {
        "upload_id": str(record.id),
        "status": record.status.value,
        "entity_type": record.detected_entity_type,
        "entity_confidence": record.entity_confidence,
        "total_rows": record.total_rows or 0,
        "valid_rows": record.valid_rows or 0,
        "invalid_rows": record.invalid_rows or 0,
        "imported_rows": record.imported_rows or 0,
        "tally_queued": (record.commit_summary or {}).get("tally_queued", 0),
        "commit_summary": record.commit_summary,
        "completed_at": record.completed_at.isoformat() if record.completed_at else None,
    }


# ─── Re-sync a completed upload to TallyPrime ────────────────────────────────

@router.post("/ingest/{upload_id}/sync-to-tally")
def ingest_resync_tally(
    upload_id: str,
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    """Queue Tally jobs for all non-error rows of a previously committed upload."""
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")
    if record.status not in (UploadStatus.COMPLETED, UploadStatus.PARTIAL):
        raise HTTPException(400, "Upload has not been committed yet")

    connector = db.query(TallyConnector).filter(
        TallyConnector.company_id == current_user.company_id,
        TallyConnector.status == ConnectorStatus.ACTIVE,
    ).first()
    if not connector:
        raise HTTPException(400, "No active TallyPrime connector found")

    from app.services.ingestion_service import VOUCHER_SUBTYPE_OP, ENTITY_TALLY_OP
    from app.models.tally_job import TallyIntegrationJob as TJI

    entity_type = record.detected_entity_type or ""
    row_report: list[dict] = record.row_report or []

    # Only resync rows that were actually committed (stored in commit_summary)
    committed_ids: set = set((record.commit_summary or {}).get("committed_row_ids", []))
    # Fall back to all non-error rows for older uploads that don't have committed_row_ids
    use_committed_filter = len(committed_ids) > 0

    batch_id = uuid.uuid4()
    queued = 0

    for row in row_report:
        row_id = row.get("row_id")
        if use_committed_filter:
            if row_id not in committed_ids:
                continue
        else:
            if row.get("status") == "error":
                continue
        mapped: dict = row.get("mapped", {})
        name = str(mapped.get("name", "")).strip()

        # Build operation + payload per entity type
        if entity_type == "Ledger":
            if not name:
                continue
            op = TallyJobOperation.CREATE_LEDGER
            payload = {
                "name": name,
                "group": str(mapped.get("parent_group") or "Sundry Debtors").strip(),
                "opening_balance": str(int(float(mapped.get("opening_balance") or 0))),
            }
        elif entity_type == "Stock Item":
            if not name:
                continue
            op = TallyJobOperation.CREATE_STOCK_ITEM
            payload = {"name": name}
            if mapped.get("stock_group"):
                payload["stock_group"] = str(mapped["stock_group"])
            if mapped.get("unit"):
                payload["unit"] = str(mapped["unit"])
            if mapped.get("rate") is not None:
                payload["rate"] = str(mapped["rate"])
        elif entity_type == "Stock Group":
            if not name:
                continue
            op = TallyJobOperation.CREATE_STOCK_GROUP
            payload = {"name": name}
            if mapped.get("parent"):
                payload["parent"] = str(mapped["parent"])
        elif entity_type == "Voucher":
            vtype = str(mapped.get("voucher_type", "")).strip().title()
            op = VOUCHER_SUBTYPE_OP.get(vtype)
            if not op:
                continue
            payload = {k: (str(v) if v is not None else None) for k, v in mapped.items() if v is not None}
        else:
            continue

        ikey = f"resync_{upload_id}_{row['row_id']}"
        if not db.query(TJI).filter(TJI.idempotency_key == ikey).first():
            db.add(TJI(
                company_id=current_user.company_id,
                connector_id=connector.id,
                operation=op,
                payload=payload,
                idempotency_key=ikey,
                batch_id=batch_id,
            ))
            queued += 1

    db.commit()
    return {"queued": queued, "batch_id": str(batch_id)}


# ─── Upload history ───────────────────────────────────────────────────────────

@router.get("")
def list_uploads(
    page: int = 1,
    page_size: int = 10,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Upload).filter(
        Upload.company_id == current_user.company_id
    ).order_by(Upload.created_at.desc())

    total = q.count()
    uploads = q.offset((page - 1) * page_size).limit(page_size).all()

    from app.models.upload_row import UploadRow, UploadRowStatus
    from sqlalchemy import func as sqlfunc

    # Row-level counts per upload
    row_counts_q = db.query(
        UploadRow.upload_id,
        UploadRow.status,
        sqlfunc.count().label("cnt"),
    ).filter(
        UploadRow.upload_id.in_([u.id for u in uploads])
    ).group_by(UploadRow.upload_id, UploadRow.status).all()

    counts: dict = {}
    for rc in row_counts_q:
        uid = str(rc.upload_id)
        counts.setdefault(uid, {})
        counts[uid][rc.status.value] = rc.cnt

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, -(-total // page_size)),
        "items": [
            {
                "id": str(u.id),
                "original_filename": u.original_filename,
                "upload_type": u.upload_type.value if u.upload_type else None,
                "status": u.status.value,
                "total_rows": u.total_rows or 0,
                "valid_rows": u.valid_rows or 0,
                "invalid_rows": u.invalid_rows or 0,
                "imported_rows": u.imported_rows or 0,
                "created_at": u.created_at.isoformat() if u.created_at else None,
                "completed_at": u.completed_at.isoformat() if u.completed_at else None,
                "errors": (u.error_summary or {}).get("errors", [])[:5],
                "row_counts": counts.get(str(u.id), {}),
            }
            for u in uploads
        ],
    }


@router.delete("/{upload_id}")
def delete_upload(
    upload_id: str,
    current_user: User = Depends(require_admin_or_accountant),
    db: Session = Depends(get_db),
):
    record = db.query(Upload).filter(
        Upload.id == upload_id,
        Upload.company_id == current_user.company_id,
    ).first()
    if not record:
        raise HTTPException(404, "Upload not found")
    db.delete(record)
    db.commit()
    return {"deleted": True}
