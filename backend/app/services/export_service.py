"""
Shared export utility — CSV, XLSX, JSON, and PDF serializers.

Usage:
    from app.services.export_service import build_export_response

    col_defs = [("Name", "name"), ("Amount", "amount"), ...]
    rows     = [{"name": "Cash", "amount": 1000}, ...]
    return build_export_response(rows, col_defs, fmt="xlsx",
                                 filename_base="ledgers",
                                 title="Ledgers Export",
                                 company_name="Acme Ltd",
                                 period_str="Jan 2026")
"""
import csv
import io
import json
import os
import uuid
from datetime import datetime, timezone
from typing import Any

from fastapi.responses import Response, StreamingResponse


ColDef = tuple[str, str]  # (display_header, dict_key)


# ─── CSV ──────────────────────────────────────────────────────────────────────

def _to_csv(rows: list[dict], col_defs: list[ColDef]) -> StreamingResponse:
    headers = [h for h, _ in col_defs]
    keys = [k for _, k in col_defs]

    def _generate():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        yield buf.getvalue()
        for row in rows:
            buf.seek(0); buf.truncate(0)
            writer.writerow([_safe_str(row.get(k)) for k in keys])
            yield buf.getvalue()

    return StreamingResponse(
        _generate(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment"},
    )


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def _to_xlsx(rows: list[dict], col_defs: list[ColDef]) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    headers = [h for h, _ in col_defs]
    keys = [k for _, k in col_defs]

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    hdr_fill = PatternFill("solid", fgColor="4F46E5")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[1].height = 18

    alt_fill = PatternFill("solid", fgColor="F8F7FF")
    for idx, row in enumerate(rows, start=2):
        ws.append([_safe_cell(row.get(k)) for k in keys])
        fill = alt_fill if idx % 2 == 0 else None
        for cell in ws[idx]:
            cell.border = border
            if fill:
                cell.fill = fill

    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value or ""))
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment"},
    )


# ─── JSON ─────────────────────────────────────────────────────────────────────

def _to_json(rows: list[dict], col_defs: list[ColDef]) -> Response:
    keys = [k for _, k in col_defs]
    headers = [h for h, _ in col_defs]
    out = [{headers[i]: _safe_str(row.get(k)) for i, k in enumerate(keys)} for row in rows]
    return Response(
        content=json.dumps(out, ensure_ascii=False, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": "attachment"},
    )


# ─── PDF ──────────────────────────────────────────────────────────────────────

def _to_pdf(
    rows: list[dict],
    col_defs: list[ColDef],
    title: str,
    company_name: str,
    period_str: str,
    upload_dir: str = "./uploads",
) -> Response:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    # Font registration (same candidates as reports.py)
    _font_candidates = [
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf", "Arial", "Arial-Bold"),
        ("C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf", "Calibri", "Calibri-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "DejaVu", "DejaVu-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Liberation", "Liberation-Bold"),
        ("/Library/Fonts/Arial.ttf", "/Library/Fonts/Arial Bold.ttf", "Arial", "Arial-Bold"),
    ]
    already = set(pdfmetrics.getRegisteredFontNames())
    FONT, FONT_BOLD = "Helvetica", "Helvetica-Bold"
    for rp, bp, rn, bn in _font_candidates:
        if os.path.exists(rp) and os.path.exists(bp):
            try:
                if rn not in already:
                    pdfmetrics.registerFont(TTFont(rn, rp))
                if bn not in already:
                    pdfmetrics.registerFont(TTFont(bn, bp))
                FONT, FONT_BOLD = rn, bn
                break
            except Exception:
                continue

    def _style(name, size, bold=False, color=None, align="LEFT"):
        return ParagraphStyle(
            name,
            fontName=FONT_BOLD if bold else FONT,
            fontSize=size,
            leading=size * 1.4,
            textColor=color or colors.HexColor("#1E293B"),
            alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 0),
        )

    os.makedirs(upload_dir, exist_ok=True)
    tmp_path = f"{upload_dir}/export_{uuid.uuid4()}.pdf"

    headers = [h for h, _ in col_defs]
    keys = [k for _, k in col_defs]
    n_cols = len(headers)

    # Auto-size columns: available width ≈ 17 cm (A4 portrait with 2 cm margins each side)
    # For wide tables, switch to landscape
    page_size = landscape(A4) if n_cols > 6 else A4
    avail_w = (page_size[0] - 4 * cm)
    col_w = avail_w / n_cols

    doc = SimpleDocTemplate(
        tmp_path,
        pagesize=page_size,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    story = []

    story.append(Paragraph("FinPilot AI", _style("brand", 20, bold=True, color=colors.HexColor("#4F46E5"))))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(company_name, _style("co", 13, bold=True)))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#4F46E5")))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(title, _style("title", 14, bold=True)))
    if period_str:
        story.append(Paragraph(period_str, _style("period", 10, color=colors.HexColor("#64748B"))))
    story.append(Paragraph(
        f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}  •  {len(rows)} record(s)",
        _style("gen", 9, color=colors.HexColor("#94A3B8")),
    ))
    story.append(Spacer(1, 0.6 * cm))

    # Table: chunk into pages of 200 rows max to keep memory reasonable
    CHUNK = 200
    for start in range(0, max(1, len(rows)), CHUNK):
        chunk = rows[start:start + CHUNK]
        table_data = [headers] + [
            [_safe_str(row.get(k)) for k in keys]
            for row in chunk
        ]
        tbl = Table(table_data, colWidths=[col_w] * n_cols, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 1), (-1, -1), FONT),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F7FF")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)
        if start + CHUNK < len(rows):
            story.append(Spacer(1, 0.4 * cm))

    story.append(Spacer(1, 1.5 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        f"Generated by FinPilot AI  •  {company_name}  •  Confidential",
        _style("footer", 8, color=colors.HexColor("#94A3B8")),
    ))

    def _set_meta(canvas, _doc):
        canvas.setTitle(title)
        canvas.setAuthor(company_name)
        canvas.setSubject("FinPilot Data Export")
        canvas.setCreator("FinPilot AI")

    doc.build(story, onFirstPage=_set_meta, onLaterPages=_set_meta)

    with open(tmp_path, "rb") as f:
        content = f.read()
    try:
        os.remove(tmp_path)
    except OSError:
        pass

    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment"},
    )


# ─── Main entry point ─────────────────────────────────────────────────────────

MIME_TYPES = {
    "csv":  "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "json": "application/json",
    "pdf":  "application/pdf",
}

EXT_MAP = {"csv": "csv", "xlsx": "xlsx", "json": "json", "pdf": "pdf"}


def build_export_response(
    rows: list[dict],
    col_defs: list[ColDef],
    fmt: str,
    filename_base: str,
    title: str,
    company_name: str,
    period_str: str = "",
    upload_dir: str = "./uploads",
) -> Response | StreamingResponse:
    fmt = fmt.lower().strip()
    if fmt not in MIME_TYPES:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Unsupported format '{fmt}'. Use csv, xlsx, json, or pdf.")

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    safe_base = filename_base.replace(" ", "_")
    filename = f"{safe_base}_{ts}.{EXT_MAP[fmt]}"

    resp: Response | StreamingResponse
    if fmt == "csv":
        resp = _to_csv(rows, col_defs)
    elif fmt == "xlsx":
        resp = _to_xlsx(rows, col_defs)
    elif fmt == "json":
        resp = _to_json(rows, col_defs)
    else:
        resp = _to_pdf(rows, col_defs, title, company_name, period_str, upload_dir=upload_dir)

    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.headers["X-Export-Rows"] = str(len(rows))
    return resp


# ─── Private helpers ──────────────────────────────────────────────────────────

def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, float):
        return f"{v:,.2f}" if v != int(v) else f"{int(v):,}"
    return str(v)


def _safe_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (int, float)):
        return v
    return str(v)
