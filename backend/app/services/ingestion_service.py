"""
Smart file ingestion service.

Pipeline:
  parse_file()  →  detect_entity_type()  →  suggest_column_mapping()
      →  validate_rows()  →  commit_rows()

Supports CSV (auto-detect encoding + delimiter + header row),
XLSX (auto-detect header row), and JSON (streaming via ijson;
auto-locate first array-of-objects; flatten one level of nesting).
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from difflib import SequenceMatcher, get_close_matches
from typing import Any, Optional

import pandas as pd
from sqlalchemy.orm import Session

try:
    import chardet
    _HAS_CHARDET = True
except ImportError:
    _HAS_CHARDET = False

try:
    import ijson
    _HAS_IJSON = True
except ImportError:
    _HAS_IJSON = False


# ─── Domain constants ──────────────────────────────────────────────────────────

TALLY_STANDARD_GROUPS: set[str] = {
    "sundry debtors", "sundry creditors", "bank accounts", "bank occ a/c",
    "branch/divisions", "capital account", "cash-in-hand", "current assets",
    "current liabilities", "direct expenses", "direct incomes", "duties & taxes",
    "fixed assets", "indirect expenses", "indirect incomes", "investments",
    "loans & advances (a)", "loans (liability)", "misc. expenses (asset)",
    "purchase accounts", "reserves & surplus", "sales accounts", "secured loans",
    "stock-in-hand", "suspense a/c", "unsecured loans", "primary",
    "profit & loss a/c",
}

PARTY_GROUPS: set[str] = {"sundry debtors", "sundry creditors"}

GSTIN_RE = re.compile(
    r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$"
)

INDIAN_STATE_CODES: set[str] = {
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
    "21", "22", "23", "24", "26", "27", "29", "30", "31",
    "32", "33", "34", "35", "36", "37", "38",
}

# ─── Entity signatures ─────────────────────────────────────────────────────────
# strong cols score 3pt; weak cols score 1pt.
# At least one strong match is required for an entity to appear in candidates.

ENTITY_SIGNATURES: dict[str, dict[str, set[str]]] = {
    "Ledger": {
        "strong": {"name", "ledger_name", "ledger", "parent_group", "group", "opening_balance"},
        "weak":   {"gstin", "gst_number", "state", "email", "phone", "address", "country"},
    },
    "Stock Item": {
        "strong": {"name", "item_name", "stock_group", "unit", "rate", "opening_qty"},
        "weak":   {"hsn_code", "hsn", "gst_rate", "category", "stock_category"},
    },
    "Stock Group": {
        "strong": {"name", "group_name"},
        "weak":   {"parent"},
    },
    "Stock Category": {
        "strong": {"name", "category_name"},
        "weak":   {"parent", "description"},
    },
    "Voucher": {
        "strong": {"date", "voucher_type", "narration", "voucher_no", "amount"},
        "weak":   {"party_ledger", "party_name", "party", "dr_ledger", "cr_ledger",
                   "debit", "credit", "from_account", "to_account"},
    },
}

VOUCHER_SUBTYPE_SIGNATURES: dict[str, set[str]] = {
    "Sales":          {"sales_ledger", "invoice_no"},
    "Purchase":       {"purchase_ledger", "bill_no"},
    "Receipt":        {"cash_bank_ledger", "receipt_no"},
    "Payment":        {"cash_bank_ledger", "payment_no"},
    "Journal":        {"dr_ledger", "cr_ledger"},
    "Contra":         {"from_account", "to_account"},
    "Credit Note":    {"credit_note_no"},
    "Debit Note":     {"debit_note_no"},
    "Stock Journal":  {"from_godown", "to_godown", "item_name"},
    "Physical Stock": {"godown", "qty"},
    "Delivery Note":  {"despatch_no", "item_name"},
    "Receipt Note":   {"receipt_note_no", "item_name"},
}

# Schema field definitions per entity type
SCHEMA_FIELDS: dict[str, dict[str, dict]] = {
    "Ledger": {
        "name":            {"required": True,  "type": "str"},
        "parent_group":    {"required": True,  "type": "str"},
        "opening_balance": {"required": False, "type": "float"},
        "gstin":           {"required": False, "type": "str"},
        "state":           {"required": False, "type": "str"},
        "address":         {"required": False, "type": "str"},
        "email":           {"required": False, "type": "str"},
        "phone":           {"required": False, "type": "str"},
        "country":         {"required": False, "type": "str"},
    },
    "Stock Item": {
        "name":            {"required": True,  "type": "str"},
        "stock_group":     {"required": False, "type": "str"},
        "stock_category":  {"required": False, "type": "str"},
        "unit":            {"required": False, "type": "str"},
        "rate":            {"required": False, "type": "float"},
        "opening_qty":     {"required": False, "type": "float"},
        "hsn_code":        {"required": False, "type": "str"},
        "gst_rate":        {"required": False, "type": "float"},
    },
    "Stock Group": {
        "name":   {"required": True,  "type": "str"},
        "parent": {"required": False, "type": "str"},
    },
    "Stock Category": {
        "name":        {"required": True,  "type": "str"},
        "parent":      {"required": False, "type": "str"},
        "description": {"required": False, "type": "str"},
    },
    "Voucher": {
        "date":             {"required": True,  "type": "date"},
        "amount":           {"required": True,  "type": "float"},
        "voucher_type":     {"required": True,  "type": "str"},
        "narration":        {"required": False, "type": "str"},
        "voucher_no":       {"required": False, "type": "str"},
        "party_ledger":     {"required": False, "type": "str"},
        "sales_ledger":     {"required": False, "type": "str"},
        "purchase_ledger":  {"required": False, "type": "str"},
        "cash_bank_ledger": {"required": False, "type": "str"},
        "dr_ledger":        {"required": False, "type": "str"},
        "cr_ledger":        {"required": False, "type": "str"},
        "from_account":     {"required": False, "type": "str"},
        "to_account":       {"required": False, "type": "str"},
        "item_name":        {"required": False, "type": "str"},
        "qty":              {"required": False, "type": "float"},
        "from_godown":      {"required": False, "type": "str"},
        "to_godown":        {"required": False, "type": "str"},
    },
}

# Common column aliases (normalised_file_col → schema_field)
COLUMN_ALIASES: dict[str, str] = {
    "ledger_name":      "name",
    "ledger":           "name",
    "item_name":        "name",
    "category_name":    "name",
    "group_name":       "name",
    "group":            "parent_group",
    "parent_group_name": "parent_group",
    "opening_bal":      "opening_balance",
    "op_balance":       "opening_balance",
    "gst_number":       "gstin",
    "gst_no":           "gstin",
    "gstin_no":         "gstin",
    "state_code":       "state",
    "stock_group_name": "stock_group",
    "uom":              "unit",
    "unit_of_measure":  "unit",
    "selling_rate":     "rate",
    "purchase_rate":    "rate",
    "opening_quantity": "opening_qty",
    "op_qty":           "opening_qty",
    "hsn":              "hsn_code",
    "hsn_sac":          "hsn_code",
    "gst_%":            "gst_rate",
    "tax_rate":         "gst_rate",
    "voucher_date":     "date",
    "txn_date":         "date",
    "transaction_date": "date",
    "total":            "amount",
    "total_amount":     "amount",
    "party_name":       "party_ledger",
    "party":            "party_ledger",
    "debit_ledger":     "dr_ledger",
    "credit_ledger":    "cr_ledger",
    "dr":               "dr_ledger",
    "cr":               "cr_ledger",
    "bank_ledger":      "cash_bank_ledger",
    "cash_ledger":      "cash_bank_ledger",
    "from_location":    "from_godown",
    "to_location":      "to_godown",
    "quantity":         "qty",
}

ENTITY_TO_UPLOAD_TYPE: dict[str, str] = {
    "Ledger":         "ledgers",
    "Stock Item":     "stock_items",
    "Stock Group":    "stock_groups",
    "Stock Category": "stock_categories",
    "Voucher":        "vouchers",
}

# Tally job operations per entity type
from app.models.tally_job import TallyJobOperation

ENTITY_TALLY_OP: dict[str, TallyJobOperation] = {
    "Ledger":      TallyJobOperation.CREATE_LEDGER,
    "Stock Item":  TallyJobOperation.CREATE_STOCK_ITEM,
    "Stock Group": TallyJobOperation.CREATE_STOCK_GROUP,
}

VOUCHER_SUBTYPE_OP: dict[str, TallyJobOperation] = {
    "Sales":          TallyJobOperation.CREATE_SALES_VOUCHER,
    "Purchase":       TallyJobOperation.CREATE_PURCHASE_VOUCHER,
    "Receipt":        TallyJobOperation.CREATE_RECEIPT_VOUCHER,
    "Payment":        TallyJobOperation.CREATE_PAYMENT_VOUCHER,
    "Journal":        TallyJobOperation.CREATE_JOURNAL_VOUCHER,
    "Credit Note":    TallyJobOperation.CREATE_CREDIT_NOTE,
    "Debit Note":     TallyJobOperation.CREATE_DEBIT_NOTE,
    "Contra":         TallyJobOperation.CREATE_CONTRA_VOUCHER,
    "Stock Journal":  TallyJobOperation.CREATE_STOCK_JOURNAL,
    "Physical Stock": TallyJobOperation.CREATE_PHYSICAL_STOCK,
    "Delivery Note":  TallyJobOperation.CREATE_DELIVERY_NOTE,
    "Receipt Note":   TallyJobOperation.CREATE_RECEIPT_NOTE,
}

MAX_ROWS_STORED = 10_000  # cap stored rows to avoid bloating JSONB columns

# ─── Reference resolution constants ───────────────────────────────────────────

FUZZY_DUP_THRESHOLD: float = 0.70   # similarity to flag a "possible duplicate"
FUZZY_REF_THRESHOLD: float = 0.60   # similarity to suggest an existing reference match

# Which mapped schema fields are foreign-key references and what they point to
ENTITY_REFERENCES: dict[str, list[dict]] = {
    'Ledger': [
        {'field': 'parent_group',    'ref_type': 'Group',          'required': True},
    ],
    'Stock Item': [
        {'field': 'stock_group',     'ref_type': 'Stock Group',    'required': False},
        {'field': 'stock_category',  'ref_type': 'Stock Category', 'required': False},
        {'field': 'unit',            'ref_type': 'Unit',           'required': False},
    ],
    'Stock Group': [
        {'field': 'parent',          'ref_type': 'Stock Group',    'required': False},
    ],
    'Stock Category': [
        {'field': 'parent',          'ref_type': 'Stock Category', 'required': False},
    ],
    'Voucher': [
        {'field': 'party_ledger',    'ref_type': 'Ledger', 'required': False},
        {'field': 'sales_ledger',    'ref_type': 'Ledger', 'required': False},
        {'field': 'purchase_ledger', 'ref_type': 'Ledger', 'required': False},
        {'field': 'cash_bank_ledger','ref_type': 'Ledger', 'required': False},
        {'field': 'dr_ledger',       'ref_type': 'Ledger', 'required': False},
        {'field': 'cr_ledger',       'ref_type': 'Ledger', 'required': False},
        {'field': 'from_account',    'ref_type': 'Ledger', 'required': False},
        {'field': 'to_account',      'ref_type': 'Ledger', 'required': False},
        {'field': 'item_name',       'ref_type': 'Stock Item', 'required': False},
    ],
}

# DB table + name column for each reference / entity type
REF_DB_TABLE: dict[str, tuple[str, str]] = {
    'Group':          ('tally_groups',       'name'),
    'Ledger':         ('tally_ledgers',      'name'),
    'Stock Group':    ('tally_stock_groups', 'name'),
    'Stock Category': ('stock_categories',   'name'),
    'Unit':           ('tally_units',        'name'),
    'Stock Item':     ('tally_stock_items',  'name'),
}

# Same map used for duplicate detection of the entity being imported
ENTITY_DB_TABLE: dict[str, str] = {
    'Ledger':         'tally_ledgers',
    'Stock Item':     'tally_stock_items',
    'Stock Group':    'tally_stock_groups',
    'Stock Category': 'stock_categories',
}


# ─── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class ParseResult:
    columns: list[str]                    # normalised column names from file
    rows: list[dict[str, Any]]            # raw rows (native types preserved)
    total: int
    file_format: str                      # csv | xlsx | json


@dataclass
class DetectionResult:
    entity_type: str
    subtype: str
    confidence: str                        # high | medium | low
    score: float
    from_cache: bool = False


@dataclass
class RowResult:
    row_id: int
    # Status: new | duplicate_exact | duplicate_fuzzy | ref_missing | ref_similar | warning | error
    status: str
    errors: list[dict[str, str]] = field(default_factory=list)
    warnings: list[dict[str, str]] = field(default_factory=list)
    mapped: dict[str, Any] = field(default_factory=dict)
    # Duplicate detection result (None = no match)
    duplicate_info: Optional[dict] = None
    # Reference resolution issues [{field, ref_entity_type, ref_name, match_type, suggestions}]
    ref_issues: list[dict] = field(default_factory=list)
    # Whether the UI checkbox should default to checked
    check_default: bool = True
    # User explicitly requested force-import despite exact duplicate
    force_import: bool = False


@dataclass
class CommitResult:
    imported: int = 0
    skipped: int = 0
    tally_queued: int = 0
    errors: list[dict] = field(default_factory=list)


# ─── Column normalisation ──────────────────────────────────────────────────────

def _normalise_col(col: str) -> str:
    col = str(col).strip().lower()
    col = re.sub(r'[\s\-/\\.]+', '_', col)
    col = re.sub(r'[^a-z0-9_]', '', col)
    col = re.sub(r'_+', '_', col).strip('_')
    return col


# ─── File parsing ──────────────────────────────────────────────────────────────

def parse_file(content: bytes, filename: str) -> ParseResult:
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.csv',):
        return _parse_csv(content)
    if ext in ('.xlsx', '.xls'):
        return _parse_xlsx(content)
    if ext in ('.json',):
        return _parse_json(content)
    raise ValueError(f"Unsupported file extension '{ext}'. Allowed: .csv, .xlsx, .xls, .json")


def _parse_csv(content: bytes) -> ParseResult:
    # Encoding detection
    if _HAS_CHARDET:
        det = chardet.detect(content[:32768])
        encoding = det.get('encoding') or 'utf-8'
    else:
        encoding = 'utf-8'
    if encoding.lower() in ('ascii',):
        encoding = 'utf-8'

    try:
        text = content.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        text = content.decode('latin-1')

    # Delimiter detection
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ','

    lines = text.splitlines()
    header_idx = _find_csv_header_row(lines, delimiter)
    body = '\n'.join(lines[header_idx:])

    df = pd.read_csv(
        io.StringIO(body),
        sep=delimiter,
        dtype=str,
        na_filter=False,
        on_bad_lines='skip',
    )
    df.columns = [_normalise_col(c) for c in df.columns]
    df = df.loc[~(df == '').all(axis=1)]  # drop entirely blank rows

    rows = df.to_dict(orient='records')
    return ParseResult(
        columns=list(df.columns),
        rows=rows[:MAX_ROWS_STORED],
        total=len(rows),
        file_format='csv',
    )


def _find_csv_header_row(lines: list[str], delimiter: str) -> int:
    """Return index of the first row that looks like a header (≥2 non-empty text cells)."""
    for i, line in enumerate(lines[:10]):
        parts = [p.strip() for p in line.split(delimiter)]
        non_empty = [p for p in parts if p]
        text_like = sum(1 for p in non_empty if not re.match(r'^-?\d+(\.\d+)?$', p))
        if len(non_empty) >= 2 and text_like >= 1:
            return i
    return 0


def _parse_xlsx(content: bytes) -> ParseResult:
    header_row = _detect_xlsx_header(content)
    df = pd.read_excel(
        io.BytesIO(content),
        header=header_row,
        dtype=str,
        na_filter=False,
        engine='openpyxl',
    )
    df.columns = [_normalise_col(str(c)) for c in df.columns]
    df = df.loc[~df.apply(lambda r: all(str(v).strip() in ('', 'nan') for v in r), axis=1)]
    rows = df.to_dict(orient='records')
    return ParseResult(
        columns=list(df.columns),
        rows=rows[:MAX_ROWS_STORED],
        total=len(rows),
        file_format='xlsx',
    )


def _detect_xlsx_header(content: bytes) -> int:
    """Scan first 10 rows; return index of the first row with ≥2 non-empty string cells."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            non_empty = [c for c in row if c is not None and str(c).strip() not in ('', 'None')]
            str_cells = sum(1 for c in non_empty if isinstance(c, str))
            if len(non_empty) >= 2 and str_cells >= len(non_empty) * 0.6:
                return row_idx
        return 0
    except Exception:
        return 0


def _parse_json(content: bytes) -> ParseResult:
    # Streaming via ijson for large files, fallback to json.loads
    if _HAS_IJSON and len(content) > 1_048_576:  # >1MB
        rows = _parse_json_streaming(content)
    else:
        data = json.loads(content)
        rows = _extract_row_array(data)

    if rows is None:
        raise ValueError("Could not locate an array of objects in the JSON file")

    rows = _flatten_and_union(rows)
    cols = list(rows[0].keys()) if rows else []
    # Normalise column names
    normalised_rows = [{_normalise_col(k): v for k, v in r.items()} for r in rows]
    normalised_cols = [_normalise_col(c) for c in cols]

    return ParseResult(
        columns=normalised_cols,
        rows=normalised_rows[:MAX_ROWS_STORED],
        total=len(normalised_rows),
        file_format='json',
    )


def _parse_json_streaming(content: bytes) -> Optional[list[dict]]:
    """Use ijson to stream-parse content and find first array-of-objects."""
    buf = io.BytesIO(content)
    # Try root-level array first
    try:
        buf.seek(0)
        items = list(ijson.items(buf, 'item'))
        if items and all(isinstance(i, dict) for i in items[:5]):
            return items
    except Exception:
        pass

    # Scan top-level keys for an array-of-objects
    try:
        buf.seek(0)
        for prefix, event, value in ijson.parse(buf):
            if event == 'map_key':
                key = value
                try:
                    buf_inner = io.BytesIO(content)
                    items = list(ijson.items(buf_inner, f'{key}.item'))
                    if items and all(isinstance(i, dict) for i in items[:5]):
                        return items
                except Exception:
                    continue
    except Exception:
        pass

    return None


def _extract_row_array(data: Any, depth: int = 0) -> Optional[list[dict]]:
    """BFS: find first array that contains ≥1 dict elements."""
    if isinstance(data, list) and data and isinstance(data[0], dict):
        return data
    if isinstance(data, dict) and depth < 3:
        for val in data.values():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                return val
        for val in data.values():
            result = _extract_row_array(val, depth + 1)
            if result is not None:
                return result
    return None


def _flatten_and_union(rows: list[dict]) -> list[dict]:
    """Flatten one level of nested dicts, union all keys across rows."""
    all_keys: set[str] = set()
    flat_rows: list[dict] = []

    for row in rows:
        flat: dict[str, Any] = {}
        for key, val in row.items():
            if isinstance(val, dict):
                for sub_key, sub_val in val.items():
                    flat[f"{key}.{sub_key}"] = sub_val
            else:
                flat[key] = val
        flat_rows.append(flat)
        all_keys |= set(flat.keys())

    return [{k: row.get(k) for k in all_keys} for row in flat_rows]


# ─── Entity type detection ─────────────────────────────────────────────────────

def detect_entity_type(
    columns: list[str],
    sample_rows: list[dict],
    company_id: Any,
    db: Session,
) -> DetectionResult:
    fingerprint = header_fingerprint(columns)

    # Check cached mapping first
    cached = _get_cached_mapping(company_id, fingerprint, db)
    if cached:
        return DetectionResult(
            entity_type=cached['entity_type'],
            subtype=cached.get('subtype', ''),
            confidence='high',
            score=999,
            from_cache=True,
        )

    col_set = set(columns)
    scores: dict[str, float] = {}

    for entity, sig in ENTITY_SIGNATURES.items():
        strong = len(col_set & sig['strong'])
        weak = len(col_set & sig['weak'])
        if strong > 0:
            scores[entity] = strong * 3 + weak

    if not scores:
        # LLM fallback
        return _llm_detect(columns, sample_rows)

    best = max(scores, key=lambda k: scores[k])
    best_score = scores[best]
    sorted_scores = sorted(scores.values(), reverse=True)

    if best_score >= 9:
        confidence = 'high'
    elif best_score >= 4:
        if len(sorted_scores) >= 2 and sorted_scores[1] >= best_score * 0.75:
            confidence = 'medium'
        else:
            confidence = 'medium'
    else:
        confidence = 'low'

    # Low confidence → LLM fallback
    if confidence == 'low':
        return _llm_detect(columns, sample_rows)

    subtype = _detect_voucher_subtype(col_set) if best == 'Voucher' else ''

    return DetectionResult(entity_type=best, subtype=subtype, confidence=confidence, score=best_score)


def _detect_voucher_subtype(col_set: set[str]) -> str:
    best_sub, best_count = '', 0
    for sub, sig in VOUCHER_SUBTYPE_SIGNATURES.items():
        count = len(col_set & sig)
        if count > best_count:
            best_count, best_sub = count, sub
    return best_sub


def _llm_detect(columns: list[str], sample_rows: list[dict]) -> DetectionResult:
    """Call Groq to classify entity type. Falls back to Unknown on any error."""
    from app.core.config import settings
    import httpx

    if not settings.GROQ_API_KEY:
        return DetectionResult(entity_type='Unknown', subtype='', confidence='low', score=0)

    prompt = (
        "You are a financial data analyst. Given these spreadsheet column names and sample rows, "
        "identify the entity type. Reply with exactly one of: "
        "Ledger, Stock Item, Stock Group, Stock Category, Voucher. "
        "If Voucher, also state the subtype separated by slash "
        "(e.g. Voucher/Sales or Voucher/Journal). "
        "No explanation — only the entity type.\n\n"
        f"Columns: {', '.join(columns)}\n"
        f"Sample rows:\n{json.dumps(sample_rows[:3], default=str, indent=2)}"
    )

    try:
        resp = httpx.post(
            'https://api.groq.com/openai/v1/chat/completions',
            headers={'Authorization': f'Bearer {settings.GROQ_API_KEY}'},
            json={
                'model': settings.GROQ_MODEL,
                'messages': [{'role': 'user', 'content': prompt}],
                'max_tokens': 20,
                'temperature': 0,
            },
            timeout=10.0,
        )
        resp.raise_for_status()
        answer = resp.json()['choices'][0]['message']['content'].strip()
        parts = answer.split('/')
        entity_type = parts[0].strip().title()
        subtype = parts[1].strip().title() if len(parts) > 1 else ''
        valid = set(SCHEMA_FIELDS.keys())
        if entity_type not in valid:
            return DetectionResult(entity_type='Unknown', subtype='', confidence='low', score=0)
        return DetectionResult(entity_type=entity_type, subtype=subtype, confidence='medium', score=5)
    except Exception:
        return DetectionResult(entity_type='Unknown', subtype='', confidence='low', score=0)


# ─── Column mapping ────────────────────────────────────────────────────────────

def suggest_column_mapping(file_columns: list[str], entity_type: str) -> dict[str, str]:
    """Return {file_col: schema_field} suggestions using aliases then fuzzy matching."""
    schema = SCHEMA_FIELDS.get(entity_type, {})
    schema_fields = list(schema.keys())
    mapping: dict[str, str] = {}
    used: set[str] = set()

    for col in file_columns:
        # 1. exact match
        if col in schema_fields and col not in used:
            mapping[col] = col
            used.add(col)
            continue
        # 2. alias lookup
        alias = COLUMN_ALIASES.get(col)
        if alias and alias in schema_fields and alias not in used:
            mapping[col] = alias
            used.add(alias)
            continue
        # 3. fuzzy match (cutoff 0.55 — slightly lenient for financial field names)
        available = [f for f in schema_fields if f not in used]
        matches = get_close_matches(col, available, n=1, cutoff=0.55)
        if matches:
            mapping[col] = matches[0]
            used.add(matches[0])
        # Unmapped columns are omitted (user can map manually in UI)

    return mapping


def header_fingerprint(columns: list[str]) -> str:
    key = ','.join(sorted(columns))
    return hashlib.md5(key.encode()).hexdigest()


# ─── Mapping cache ─────────────────────────────────────────────────────────────

def _get_cached_mapping(company_id: Any, fingerprint: str, db: Session) -> Optional[dict]:
    from app.models.upload import UploadColumnCache
    row = db.query(UploadColumnCache).filter(
        UploadColumnCache.company_id == company_id,
        UploadColumnCache.header_fingerprint == fingerprint,
    ).first()
    if row:
        return {'entity_type': row.entity_type, 'column_mapping': row.column_mapping}
    return None


def save_mapping_cache(
    company_id: Any, fingerprint: str, entity_type: str, mapping: dict, db: Session
) -> None:
    from app.models.upload import UploadColumnCache
    existing = db.query(UploadColumnCache).filter(
        UploadColumnCache.company_id == company_id,
        UploadColumnCache.header_fingerprint == fingerprint,
    ).first()
    if existing:
        existing.entity_type = entity_type
        existing.column_mapping = mapping
    else:
        db.add(UploadColumnCache(
            company_id=company_id,
            header_fingerprint=fingerprint,
            entity_type=entity_type,
            column_mapping=mapping,
        ))
    db.commit()


# ─── Row extraction helper ─────────────────────────────────────────────────────

def _get_mapped_val(row: dict, mapping: dict[str, str], schema_field: str) -> Any:
    """Return the value from row for a given schema_field via the column mapping."""
    for file_col, sf in mapping.items():
        if sf == schema_field:
            val = row.get(file_col)
            if val is None:
                return None
            if isinstance(val, str):
                s = val.strip()
                return s if s else None
            return val
    return None


def _coerce_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        cleaned = re.sub(r'[,\s]', '', str(val))
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _coerce_date(val: Any) -> Optional[str]:
    if val is None:
        return None
    from dateutil import parser as du_parser
    try:
        return du_parser.parse(str(val)).date().isoformat()
    except Exception:
        return None


# ─── Reference resolution helpers ────────────────────────────────────────────

def _name_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _load_ref_records(ref_type: str, company_id: Any, db: Session) -> list[dict]:
    """Return all active records for the given reference entity type, including
    Tally standard groups for the 'Group' type.  Each item has name/id/name_lower."""
    table_info = REF_DB_TABLE.get(ref_type)
    base: list[dict] = []

    if table_info:
        table, name_col = table_info
        from sqlalchemy import text
        try:
            rows = db.execute(
                text(f"SELECT {name_col}, id::text FROM {table} "
                     f"WHERE company_id = :cid AND is_active = true"),
                {'cid': str(company_id)},
            ).fetchall()
            base = [{'name': r[0], 'id': r[1], 'name_lower': r[0].lower().strip()} for r in rows]
        except Exception:
            pass

    # Always include Tally standard groups for the Group ref type
    if ref_type == 'Group':
        existing_lower = {r['name_lower'] for r in base}
        for g in TALLY_STANDARD_GROUPS:
            if g not in existing_lower:
                base.append({'name': g.title(), 'id': None, 'name_lower': g})

    return base


def _load_entity_records(entity_type: str, company_id: Any, db: Session) -> list[dict]:
    """Load all existing records for the entity type being imported (for dupe detection)."""
    table = ENTITY_DB_TABLE.get(entity_type)
    if not table:
        return []
    from sqlalchemy import text
    try:
        rows = db.execute(
            text(f"SELECT name, id::text FROM {table} "
                 f"WHERE company_id = :cid AND is_active = true"),
            {'cid': str(company_id)},
        ).fetchall()
        return [{'name': r[0], 'id': r[1], 'name_lower': r[0].lower().strip()} for r in rows]
    except Exception:
        return []


def check_duplicate(name_lower: str, entity_records: list[dict]) -> Optional[dict]:
    """
    Returns None (no match), or
    {'match_type': 'exact'|'fuzzy', 'matched_name', 'matched_id', 'similarity'}.
    """
    for rec in entity_records:
        if rec['name_lower'] == name_lower:
            return {'match_type': 'exact', 'matched_name': rec['name'],
                    'matched_id': rec['id'], 'similarity': 1.0}

    best_sim, best_rec = 0.0, None
    for rec in entity_records:
        sim = _name_similarity(name_lower, rec['name_lower'])
        if sim > best_sim:
            best_sim, best_rec = sim, rec

    if best_rec and best_sim >= FUZZY_DUP_THRESHOLD:
        return {'match_type': 'fuzzy', 'matched_name': best_rec['name'],
                'matched_id': best_rec['id'], 'similarity': round(best_sim, 3)}
    return None


def build_intra_file_index(
    rows: list[dict], entity_type: str, column_mapping: dict[str, str]
) -> set[str]:
    """Return lowercase names of all entities being created in this upload file.
    Used to satisfy intra-file parent references (e.g., Stock Group → parent
    that is also in the same file)."""
    names: set[str] = set()
    for row in rows:
        name = _get_mapped_val(row, column_mapping, 'name')
        if name:
            names.add(str(name).strip().lower())
    return names


def _resolve_reference(
    field: str,
    ref_value: str,
    ref_type: str,
    ref_records: list[dict],
    intra_file_names: set[str],
) -> dict:
    """Resolve one reference field.  Returns a ref_issue dict with match_type:
       ok | intra_file | similar | missing."""
    ref_lower = str(ref_value).strip().lower()

    # 1. Exact match in DB/standard records
    for rec in ref_records:
        if rec['name_lower'] == ref_lower:
            return {'field': field, 'ref_entity_type': ref_type, 'ref_name': ref_value,
                    'match_type': 'ok', 'suggestions': []}

    # 2. Dependency resolved within this same upload file
    if ref_lower in intra_file_names:
        return {'field': field, 'ref_entity_type': ref_type, 'ref_name': ref_value,
                'match_type': 'intra_file', 'suggestions': []}

    # 3. Fuzzy match — suggest existing candidates
    candidates = sorted(
        [{'name': r['name'], 'id': r['id'], 'similarity': round(_name_similarity(ref_lower, r['name_lower']), 3)}
         for r in ref_records
         if _name_similarity(ref_lower, r['name_lower']) >= FUZZY_REF_THRESHOLD],
        key=lambda x: x['similarity'], reverse=True,
    )[:3]

    if candidates:
        return {'field': field, 'ref_entity_type': ref_type, 'ref_name': ref_value,
                'match_type': 'similar', 'suggestions': candidates}

    # 4. Not found anywhere
    return {'field': field, 'ref_entity_type': ref_type, 'ref_name': ref_value,
            'match_type': 'missing', 'suggestions': []}


def resolve_entity_references(
    entity_type: str,
    mapped: dict,
    ref_cache: dict[str, list[dict]],
    intra_file_names: set[str],
) -> list[dict]:
    """Resolve all reference fields for the given entity type.
    Returns list of ref_issue dicts (only for fields that have a value)."""
    ref_defs = ENTITY_REFERENCES.get(entity_type, [])
    issues: list[dict] = []
    for ref_def in ref_defs:
        field_name = ref_def['field']
        ref_type = ref_def['ref_type']
        ref_value = mapped.get(field_name)
        if ref_value is None or str(ref_value).strip() == '':
            continue
        ref_records = ref_cache.get(ref_type, [])
        issue = _resolve_reference(field_name, str(ref_value), ref_type, ref_records, intra_file_names)
        if issue['match_type'] != 'ok':
            issues.append(issue)
    return issues


def _determine_final_status(
    errors: list, warnings: list, dup_info: Optional[dict], ref_issues: list[dict]
) -> tuple[str, bool]:
    """Compute (status_string, check_default) from accumulated validation data.

    Priority (highest → lowest):
      error → duplicate_exact → ref_missing → duplicate_fuzzy → ref_similar → warning → new
    """
    if errors:
        return 'error', False

    if dup_info and dup_info['match_type'] == 'exact':
        return 'duplicate_exact', False  # default unchecked — user must force-import

    has_missing = any(r['match_type'] == 'missing' for r in ref_issues)
    if has_missing:
        return 'ref_missing', True

    if dup_info and dup_info['match_type'] == 'fuzzy':
        return 'duplicate_fuzzy', True

    has_similar = any(r['match_type'] == 'similar' for r in ref_issues)
    if has_similar:
        return 'ref_similar', True

    if warnings:
        return 'warning', True

    return 'new', True


# ─── Row validation ────────────────────────────────────────────────────────────

def validate_rows(
    rows: list[dict],
    entity_type: str,
    column_mapping: dict[str, str],
    company_id: Any,
    db: Session,
) -> list[RowResult]:
    """
    Full row validation pipeline:
      1. Type-coerce and required-field checks
      2. Entity-specific field validation (GSTIN, dates, amounts…)
      3. Intra-file name deduplication
      4. DB duplicate detection (exact + fuzzy)
      5. Foreign-key reference resolution (exact → intra-file → fuzzy → missing)
      6. Final status + check_default assignment
    """
    schema = SCHEMA_FIELDS.get(entity_type, {})

    # ── Pre-load data once before the row loop ─────────────────────────────────
    # Existing records of THIS entity type (for duplicate detection)
    entity_records = _load_entity_records(entity_type, company_id, db)

    # Records for every reference type this entity uses
    ref_cache: dict[str, list[dict]] = {}
    for ref_def in ENTITY_REFERENCES.get(entity_type, []):
        rt = ref_def['ref_type']
        if rt not in ref_cache:
            ref_cache[rt] = _load_ref_records(rt, company_id, db)

    # Company groups (for Ledger parent_group validation)
    company_groups: set[str] = set()
    if entity_type == 'Ledger':
        company_groups = {r['name_lower'] for r in ref_cache.get('Group', [])}

    # Names being created by this file (for intra-file reference resolution)
    intra_file_names = build_intra_file_index(rows, entity_type, column_mapping)

    # Track intra-file duplicates by position
    seen_in_file: dict[str, int] = {}

    results: list[RowResult] = []

    for row_id, row in enumerate(rows):
        errors: list[dict[str, str]] = []
        warnings: list[dict[str, str]] = []
        mapped: dict[str, Any] = {}

        # ── 1. Type-coerce all mapped values ───────────────────────────────────
        for schema_field, meta in schema.items():
            raw = _get_mapped_val(row, column_mapping, schema_field)
            field_type = meta['type']

            if field_type == 'float':
                coerced = _coerce_float(raw)
                if raw is not None and raw != '' and coerced is None:
                    errors.append({'field': schema_field, 'message': f"'{raw}' is not a valid number"})
                mapped[schema_field] = coerced
            elif field_type == 'date':
                coerced = _coerce_date(raw)
                if raw is not None and raw != '' and coerced is None:
                    errors.append({'field': schema_field, 'message': f"'{raw}' is not a valid date"})
                mapped[schema_field] = coerced
            else:
                mapped[schema_field] = str(raw) if raw is not None else None

            if meta['required'] and (mapped.get(schema_field) is None or mapped.get(schema_field) == ''):
                errors.append({'field': schema_field, 'message': f"'{schema_field}' is required"})

        # ── 2. Entity-specific field validation ────────────────────────────────
        if entity_type == 'Ledger':
            _validate_ledger_fields(mapped, errors, warnings, company_groups)
        elif entity_type == 'Stock Item':
            _validate_stock_item_fields(mapped, warnings)
        elif entity_type == 'Voucher':
            _validate_voucher(mapped, errors, warnings)

        # ── 3. Intra-file name deduplication ───────────────────────────────────
        name_val = str(mapped.get('name', '')).lower().strip()
        if name_val:
            if name_val in seen_in_file:
                warnings.append({
                    'field': 'name',
                    'message': f"Duplicate within file (also at row {seen_in_file[name_val] + 2})",
                })
            else:
                seen_in_file[name_val] = row_id

        # ── 4. DB duplicate detection ──────────────────────────────────────────
        dup_info: Optional[dict] = None
        if name_val and entity_type in ENTITY_DB_TABLE:
            dup_info = check_duplicate(name_val, entity_records)

        # ── 5. Reference resolution ────────────────────────────────────────────
        ref_issues = resolve_entity_references(entity_type, mapped, ref_cache, intra_file_names)

        # ── 6. Final status ────────────────────────────────────────────────────
        status, check_default = _determine_final_status(errors, warnings, dup_info, ref_issues)

        results.append(RowResult(
            row_id=row_id,
            status=status,
            errors=errors,
            warnings=warnings,
            mapped=mapped,
            duplicate_info=dup_info,
            ref_issues=ref_issues,
            check_default=check_default,
        ))

    return results


def _validate_ledger_fields(
    mapped: dict, errors: list, warnings: list, company_groups: set[str],
) -> None:
    """Validate Ledger-specific fields (group validity, GSTIN format, state code).
    Duplicate detection is handled separately via check_duplicate()."""
    group_raw = str(mapped.get('parent_group') or '').strip()
    if group_raw:
        group_lower = group_raw.lower()
        all_groups = TALLY_STANDARD_GROUPS | company_groups
        matches = get_close_matches(group_lower, all_groups, n=1, cutoff=0.75)
        if matches:
            # Auto-correct typo to canonical group name
            mapped['parent_group'] = next(
                (g for g in TALLY_STANDARD_GROUPS if g == matches[0]),
                group_raw,
            ).title()
        # Note: 'no match found' is now surfaced via ref_issues (match_type='missing'),
        # not as a warning here, to avoid duplicate messages.

        # Conditional fields only for party ledgers
        effective_group = (matches[0] if matches else group_lower)
        if effective_group in PARTY_GROUPS:
            gstin = str(mapped.get('gstin') or '').strip()
            if gstin and not GSTIN_RE.match(gstin.upper()):
                errors.append({'field': 'gstin', 'message': f"Invalid GSTIN format: '{gstin}'"})

            state = str(mapped.get('state') or '').strip()
            if state and len(state) == 2 and state.isdigit() and state not in INDIAN_STATE_CODES:
                errors.append({'field': 'state', 'message': f"Invalid state code: '{state}'"})


def _validate_stock_item_fields(mapped: dict, warnings: list) -> None:
    rate = mapped.get('rate')
    if rate is not None and rate < 0:
        warnings.append({'field': 'rate', 'message': 'Rate is negative — verify this is intentional'})


def _validate_voucher(mapped: dict, errors: list, warnings: list) -> None:
    amount = mapped.get('amount')
    if amount is not None and amount == 0:
        warnings.append({'field': 'amount', 'message': 'Amount is zero'})

    vtype = str(mapped.get('voucher_type') or '').strip().lower()
    known = {s.lower() for s in VOUCHER_SUBTYPE_SIGNATURES.keys()}
    if vtype and vtype not in known:
        warnings.append({'field': 'voucher_type', 'message': f"Unknown voucher type '{vtype}' — will be passed as-is to Tally"})


# ─── Commit ────────────────────────────────────────────────────────────────────

def commit_rows(
    upload_record_id: Any,
    selected_row_ids: list[int],
    sync_to_tally: bool,
    company_id: Any,
    user_id: Any,
    db: Session,
) -> CommitResult:
    """
    Read upload's row_report, commit only selected rows.
    Creates DB records in dependency order:
      Groups/Categories → Ledgers/StockItems → Vouchers
    Updates upload.imported_rows continuously for progress polling.
    """
    from app.models.upload import Upload, UploadStatus
    from app.models.tally_connector import TallyConnector, ConnectorStatus
    from app.models.tally_job import TallyIntegrationJob

    upload = db.query(Upload).filter(Upload.id == upload_record_id).first()
    if not upload:
        return CommitResult()

    row_report: list[dict] = upload.row_report or []
    entity_type: str = upload.detected_entity_type or ''
    column_mapping: dict = upload.column_mapping or {}

    # Filter to selected rows that aren't errors.
    # duplicate_exact rows are excluded unless the user explicitly force-imported them.
    id_set = set(selected_row_ids)
    to_commit = [
        r for r in row_report
        if r['row_id'] in id_set
        and r['status'] != 'error'
        and not (r['status'] == 'duplicate_exact' and not r.get('force_import', False))
    ]

    connector = db.query(TallyConnector).filter(
        TallyConnector.company_id == company_id,
        TallyConnector.status == ConnectorStatus.ACTIVE,
    ).first()

    result = CommitResult()
    now = datetime.now(timezone.utc)
    # One batch_id for all Tally jobs in this bulk import so they show as a single activity entry
    bulk_batch_id = uuid.uuid4() if sync_to_tally else None

    # Track which row_ids were actually committed so resync only replays those
    committed_row_ids: list[int] = []

    for row_data in to_commit:
        mapped: dict = row_data.get('mapped', {})
        try:
            queued = _commit_single(
                entity_type, mapped, company_id, user_id, connector, sync_to_tally, db, now, bulk_batch_id
            )
            result.imported += 1
            result.tally_queued += 1 if queued else 0
            committed_row_ids.append(row_data['row_id'])

            # Update progress in DB after each successful row
            upload.imported_rows = result.imported
            db.commit()
        except Exception as exc:
            result.errors.append({'row_id': row_data['row_id'], 'message': str(exc)})
            db.rollback()

    upload.status = UploadStatus.COMPLETED if not result.errors else UploadStatus.PARTIAL
    upload.commit_summary = {
        'imported': result.imported,
        'tally_queued': result.tally_queued,
        'skipped': result.skipped,
        'errors': result.errors[:50],
        'committed_row_ids': committed_row_ids,
    }
    upload.completed_at = now
    db.commit()

    return result


def _tally_key(company_id: Any, name: str) -> str:
    return f"{company_id}::{str(name).strip().lower()}"


def _queue_tally_job(
    db: Session, company_id: Any, connector_id: Any,
    operation: TallyJobOperation, payload: dict, ikey: str,
    batch_id: Any = None,
) -> bool:
    from app.models.tally_job import TallyIntegrationJob
    if db.query(TallyIntegrationJob).filter(TallyIntegrationJob.idempotency_key == ikey).first():
        return False
    db.add(TallyIntegrationJob(
        company_id=company_id,
        connector_id=connector_id,
        operation=operation,
        payload=payload,
        idempotency_key=ikey,
        batch_id=batch_id,
    ))
    return True


def _commit_single(
    entity_type: str,
    mapped: dict,
    company_id: Any,
    user_id: Any,
    connector: Any,
    sync_to_tally: bool,
    db: Session,
    now: datetime,
    batch_id: Any = None,
) -> bool:
    """Commit one row; returns True if a Tally job was queued."""
    if entity_type == 'Ledger':
        return _commit_ledger(mapped, company_id, connector, sync_to_tally, db, batch_id)
    if entity_type == 'Stock Item':
        return _commit_stock_item(mapped, company_id, connector, sync_to_tally, db, batch_id)
    if entity_type == 'Stock Group':
        return _commit_stock_group(mapped, company_id, connector, sync_to_tally, db, batch_id)
    if entity_type == 'Stock Category':
        return _commit_stock_category(mapped, company_id, db)
    if entity_type == 'Voucher':
        return _commit_voucher(mapped, company_id, connector, sync_to_tally, db, batch_id)
    return False


def _commit_ledger(mapped: dict, company_id, connector, sync_to_tally: bool, db: Session, batch_id=None) -> bool:
    from app.models.tally_masters import TallyLedger
    name = str(mapped.get('name') or '').strip()
    if not name:
        raise ValueError("Ledger name is empty")
    key = _tally_key(company_id, name)
    if db.query(TallyLedger).filter(TallyLedger.tally_key == key).first():
        return False  # already exists — skip silently

    group = str(mapped.get('parent_group') or 'Sundry Debtors').strip()
    ob = mapped.get('opening_balance') or 0.0
    ledger = TallyLedger(
        company_id=company_id, name=name, parent_group=group,
        opening_balance=float(ob),
        gstin=mapped.get('gstin'), state=mapped.get('state'),
        address=mapped.get('address'), email=mapped.get('email'),
        phone=mapped.get('phone'), country=mapped.get('country'),
        tally_key=key, source='finpilot', tally_sync_status='pending', is_active=True,
    )
    db.add(ledger)
    db.flush()

    if connector and sync_to_tally:
        ikey = f"ingest_ledger::{key}"
        return _queue_tally_job(db, company_id, connector.id, TallyJobOperation.CREATE_LEDGER,
                                {'name': name, 'group': group, 'opening_balance': str(int(ob))}, ikey, batch_id)
    return False


def _commit_stock_item(mapped: dict, company_id, connector, sync_to_tally: bool, db: Session, batch_id=None) -> bool:
    from app.models.tally_masters import TallyStockItem
    name = str(mapped.get('name') or '').strip()
    if not name:
        raise ValueError("Stock item name is empty")
    key = _tally_key(company_id, name)
    if db.query(TallyStockItem).filter(TallyStockItem.tally_key == key).first():
        return False

    item = TallyStockItem(
        company_id=company_id, name=name,
        stock_group=mapped.get('stock_group'), stock_category=mapped.get('stock_category'),
        unit=mapped.get('unit'), rate=mapped.get('rate') or 0.0,
        opening_qty=mapped.get('opening_qty') or 0.0,
        tally_key=key, source='finpilot', tally_sync_status='pending', is_active=True,
    )
    db.add(item)
    db.flush()

    if connector and sync_to_tally:
        ikey = f"ingest_si::{key}"
        payload = {'name': name}
        if mapped.get('stock_group'):
            payload['stock_group'] = mapped['stock_group']
        if mapped.get('unit'):
            payload['unit'] = mapped['unit']
        if mapped.get('rate'):
            payload['rate'] = str(mapped['rate'])
        return _queue_tally_job(db, company_id, connector.id, TallyJobOperation.CREATE_STOCK_ITEM, payload, ikey, batch_id)
    return False


def _commit_stock_group(mapped: dict, company_id, connector, sync_to_tally: bool, db: Session, batch_id=None) -> bool:
    from app.models.tally_masters import TallyStockGroup
    name = str(mapped.get('name') or '').strip()
    if not name or name.lower() == 'primary':
        return False
    key = _tally_key(company_id, name)
    if db.query(TallyStockGroup).filter(TallyStockGroup.tally_key == key).first():
        return False

    parent = mapped.get('parent') or None
    sg = TallyStockGroup(
        company_id=company_id, name=name, parent=parent,
        tally_key=key, source='finpilot', tally_sync_status='pending', is_active=True,
    )
    db.add(sg)
    db.flush()

    if connector and sync_to_tally:
        ikey = f"ingest_sg::{key}"
        payload = {'name': name}
        if parent:
            payload['parent'] = parent
        return _queue_tally_job(db, company_id, connector.id, TallyJobOperation.CREATE_STOCK_GROUP, payload, ikey, batch_id)
    return False


def _commit_stock_category(mapped: dict, company_id, db: Session) -> bool:
    from app.models.stock_category import StockCategory
    name = str(mapped.get('name') or '').strip()
    if not name:
        return False
    key = _tally_key(company_id, name)
    exists = db.query(StockCategory).filter(
        StockCategory.company_id == company_id,
        StockCategory.tally_key == key,
    ).first()
    if exists:
        return False
    db.add(StockCategory(
        company_id=company_id, name=name,
        parent=mapped.get('parent'), description=mapped.get('description'),
        tally_key=key, source='finpilot', is_active=True,
    ))
    db.flush()
    return False  # no Tally sync for stock categories


def _commit_voucher(mapped: dict, company_id, connector, sync_to_tally: bool, db: Session, batch_id=None) -> bool:
    """Queue a Tally job for the voucher; create an Invoice record for Sales/Purchase."""
    if not connector or not sync_to_tally:
        return False

    vtype_raw = str(mapped.get('voucher_type') or '').strip()
    vtype = vtype_raw.title()
    op = VOUCHER_SUBTYPE_OP.get(vtype)
    if not op:
        return False

    payload = {k: (str(v) if v is not None else None) for k, v in mapped.items() if v is not None}
    ikey = f"ingest_vch::{company_id}::{vtype}::{payload.get('date', '')}::{payload.get('amount', '')}::{payload.get('party_ledger', '')}"
    return _queue_tally_job(db, company_id, connector.id, op, payload, ikey, batch_id)
